"""
批量银行报盘合并与工资表打印工具
"""

import os
import shutil
import tkinter as tk
from datetime import datetime
from decimal import Decimal, ROUND_DOWN, ROUND_HALF_UP
from tkinter import filedialog, messagebox, scrolledtext

import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
import xlrd

HEADERS = ["序号", "账户", "户名", "金额", "跨行标识", "行名", "联行行号", "摘要", "备注"]


def split_filename(filename):
    """
    解析文件名: YYYYMM-银行名称-单位名.xls
    返回 (yearmon, bank_name, unit_name)
    注意: 单位名可能含 '-'，如 (吉林大学）深部探测与成像全国重点实验室
    策略: 第一个 '-' 前是年月，第二个 '-' 前是银行名，之后全是单位名
    """
    name_no_ext = filename.rsplit(".", 1)[0]
    parts = name_no_ext.split("-", 2)
    if len(parts) != 3:
        raise ValueError(f"文件名格式不匹配: {filename}")
    return parts[0], parts[1], parts[2]


def detect_bank_type(filename):
    if "建设银行" in filename:
        return "ccb"
    if "工商银行" in filename:
        return "icbc"
    if "吉林银行" in filename:
        return "jlb"
    known_others = ["中国银行", "交通银行", "农业银行", "九台农商行", "春城农商行"]
    for kb in known_others:
        if kb in filename:
            return "ccb"
    raise ValueError(f"无法识别银行类型: {filename}")


def rename_bank_files(bank_dir, output_dir):
    """
    扫描 bank_dir 下所有 .xls 文件
    改名规则: 单位名-YYYYMM-银行名称.xls
    复制到 output_dir（不修改源文件）
    返回: [(new_name, old_name, bank_name, unit_name), ...]
    """
    os.makedirs(output_dir, exist_ok=True)
    result = []

    for fname in os.listdir(bank_dir):
        if not fname.lower().endswith(".xls"):
            continue

        old_path = os.path.join(bank_dir, fname)
        yearmon, bank_name, unit_name = split_filename(fname)

        # 新文件名: 单位名-YYYYMM-银行名称.xls
        new_name = f"{unit_name}-{yearmon}-{bank_name}.xls"
        new_path = os.path.join(output_dir, new_name)

        shutil.copy2(old_path, new_path)
        result.append((new_name, fname, bank_name, unit_name))

    return result


def _to_decimal(val):
    """将 xlrd 读取的值转为 Decimal（两位小数，分以下直接舍去），避免 float 精度问题"""
    try:
        if isinstance(val, str):
            val = val.replace(",", "").strip()
        d = Decimal(str(val))
        # 检查是否有分以下的数值（厘）
        truncated = d.quantize(Decimal("0.01"), rounding=ROUND_DOWN)
        if truncated != d:
            return truncated, True  # (值, 有警告)
        return truncated, False
    except Exception:
        return Decimal("0.00"), False


def _read_icbc_rows(filepath, warnings):
    """读取工商银行 3 列格式，映射为建设银行 9 列"""
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    rows = []
    for r in range(1, ws.nrows):  # 跳过表头
        seq = len(rows) + 1
        account = str(ws.cell_value(r, 0)).strip()
        name = str(ws.cell_value(r, 1)).strip()
        amount, warned = _to_decimal(ws.cell_value(r, 2))
        if warned:
            warnings.append(f"{os.path.basename(filepath)} 第{r+1}行 {name} 金额 {ws.cell_value(r, 2)} 含分以下数值，已舍去")
        rows.append([seq, account, name, amount, "", "", "", "", ""])
    return rows


def _read_ccb_rows(filepath, warnings):
    """读取建设银行 9 列格式，跨行标识和行名不填，建行系统自动识别"""
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    rows = []
    for r in range(1, ws.nrows):  # 跳过表头
        seq = len(rows) + 1
        row = []
        for c in range(9):
            val = ws.cell_value(r, c)
            # 金额列（索引3）转数字
            if c == 3:
                val, warned = _to_decimal(val)
                if warned:
                    name = str(ws.cell_value(r, 2)).strip()
                    warnings.append(f"{os.path.basename(filepath)} 第{r+1}行 {name} 金额 {ws.cell_value(r, 3)} 含分以下数值，已舍去")
            row.append(val)
        # 序号重新生成
        row[0] = seq
        # 跨行标识和行名不填，建行系统自动识别
        row[4] = ""
        row[5] = ""
        rows.append(row)
    return rows


def _read_jlb_rows(filepath, warnings):
    """读取吉林银行 9 列格式，跨行标识和行名不填，建行系统自动识别"""
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    rows = []
    for r in range(1, ws.nrows):  # 跳过表头
        seq = len(rows) + 1
        row = []
        for c in range(9):
            val = ws.cell_value(r, c)
            # 金额列（索引3）转数字
            if c == 3:
                val, warned = _to_decimal(val)
                if warned:
                    name = str(ws.cell_value(r, 2)).strip()
                    warnings.append(f"{os.path.basename(filepath)} 第{r+1}行 {name} 金额 {ws.cell_value(r, 3)} 含分以下数值，已舍去")
            row.append(val)
        # 序号重新生成
        row[0] = seq
        # 跨行标识和行名不填，建行系统自动识别
        row[4] = ""
        row[5] = ""
        rows.append(row)
    return rows


def _bank_sort_key(bank_name):
    """银行排序：建设银行 < 工商银行 < 吉林银行"""
    order = {"建设银行": 0, "工商银行": 1, "吉林银行": 2}
    return order.get(bank_name, 99)


def merge_bank_files(renamed_list, bank_dir, output_dir):
    """
    按单位名分组，每个单位一个合并文件
    建设银行/吉林银行: 直接复制9列
    工商银行: 3列 → 映射到9列
    合并文件名: {单位名}-{年月}-{银行1}{行数}-{银行2}{行数}.xlsx
    银行按字母排序（建设银行 < 工商银行 < 吉林银行）
    返回: (merged_files, warnings) 已排序
    """
    os.makedirs(output_dir, exist_ok=True)

    # 按单位名分组
    units = {}
    for new_name, old_name, bank_name, unit_name in renamed_list:
        units.setdefault(unit_name, []).append((new_name, old_name, bank_name))

    merged_files = []
    warnings = []

    # 计算序号位数（根据文件总数）
    total_units = len(units)
    pad_width = len(str(total_units))

    for idx, (unit_name, entries) in enumerate(sorted(units.items()), start=1):
        # 按银行分组
        bank_groups = {}
        yearmon = None
        for new_name, old_name, bank_name in entries:
            bank_groups.setdefault(bank_name, []).append(old_name)
            # 从第一个 entry 获取年份
            if yearmon is None:
                y, _, _ = split_filename(old_name)
                yearmon = y

        # 读取每个银行的数据
        all_rows = []
        bank_summary = []
        for bank_name in sorted(bank_groups, key=_bank_sort_key):
            file_list = bank_groups[bank_name]
            bank_rows = []
            for old_name in file_list:
                filepath = os.path.join(bank_dir, old_name)
                bt = detect_bank_type(old_name)
                if bt == "icbc":
                    bank_rows.extend(_read_icbc_rows(filepath, warnings))
                elif bt == "ccb":
                    bank_rows.extend(_read_ccb_rows(filepath, warnings))
                elif bt == "jlb":
                    bank_rows.extend(_read_jlb_rows(filepath, warnings))
            # 重新编号（跨银行连续递增）
            for i, row in enumerate(bank_rows):
                row[0] = len(all_rows) + i + 1
            all_rows.extend(bank_rows)
            bank_summary.append((bank_name, len(bank_rows)))

        # 构建合并文件名（加序号前缀 + 总计金额后缀）
        seq_prefix = str(idx).zfill(pad_width)
        parts = [f"{seq_prefix}_{unit_name}", yearmon]
        for bank_name, count in bank_summary:
            parts.append(f"{bank_name}{count}个")

        # 计算总计金额（元角分，Decimal 精确计算）
        total_fen = 0
        for row in all_rows:
            amt = row[3]
            if isinstance(amt, Decimal):
                total_fen += int(amt * 100)
            else:
                total_fen += int(round(float(amt) * 100))
        yuan = total_fen // 100
        jiao = (total_fen % 100) // 10
        fen = total_fen % 10

        merged_name = "-".join(parts) + f"-总计{yuan}元{jiao}角{fen}分.xlsx"
        merged_path = os.path.join(output_dir, merged_name)

        # 写入（建行系统兼容：空值列不写入单元格，金额列设小数格式）
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = unit_name
        ws.append(HEADERS)
        for row_idx, row in enumerate(all_rows, start=2):
            for c, val in enumerate(row, start=1):
                if val is None or val == "":
                    continue  # 跳过空值，建行系统不接受空 inlineStr 单元格
                cell = ws.cell(row=row_idx, column=c, value=val)
                if c == 4:  # 金额列设两位小数格式
                    cell.number_format = "0.00"
        wb.save(merged_path)
        merged_files.append(merged_name)

    return sorted(merged_files), warnings


# ──────────────────────────────────────────────
# 银行报盘高级合并（按大单位/银行/月份合并）
# ──────────────────────────────────────────────

# ── 映射规则：从 6月银行导入模板分类清单.xlsx 的「映射规则」sheet 读取 ──
_MAPPING_EXCEL = None
_MAPPING_RULES_CACHE = None


def _get_mapping_excel_path():
    global _MAPPING_EXCEL
    if _MAPPING_EXCEL is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        _MAPPING_EXCEL = os.path.join(script_dir, "template", "映射规则.xlsx")
    return _MAPPING_EXCEL


def _load_mapping_rules():
    global _MAPPING_RULES_CACHE
    if _MAPPING_RULES_CACHE is not None:
        return _MAPPING_RULES_CACHE
    import openpyxl
    excel_path = _get_mapping_excel_path()
    if not os.path.exists(excel_path):
        _MAPPING_RULES_CACHE = []
        return _MAPPING_RULES_CACHE
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception:
        _MAPPING_RULES_CACHE = []
        return _MAPPING_RULES_CACHE
    if "映射规则" not in wb.sheetnames:
        wb.close()
        _MAPPING_RULES_CACHE = []
        return _MAPPING_RULES_CACHE
    ws = wb["映射规则"]
    rules = []
    for r in range(2, ws.max_row + 1):
        match_type = ws.cell(r, 2).value
        pattern = ws.cell(r, 3).value
        big_org = ws.cell(r, 4).value
        excluded = ws.cell(r, 5).value
        if not pattern or not match_type:
            continue
        rules.append((
            str(match_type).strip(),
            str(pattern).strip(),
            str(big_org).strip() if big_org else "",
            str(excluded).strip() == "Y",
        ))
    wb.close()
    _MAPPING_RULES_CACHE = rules
    return rules


def is_excluded(unit_name):
    rules = _load_mapping_rules()
    for match_type, pattern, _big_org, excluded in rules:
        if not excluded:
            continue
        if match_type == "prefix" and unit_name.startswith(pattern):
            return True
        if match_type == "exact" and unit_name == pattern:
            return True
        if match_type == "contains" and pattern in unit_name:
            return True
    return False


def get_big_org(unit_name):
    rules = _load_mapping_rules()
    if is_excluded(unit_name):
        return unit_name, unit_name
    for match_type, pattern, big_org, excluded in rules:
        if excluded:
            continue
        if match_type == "prefix" and unit_name.startswith(pattern):
            return big_org, unit_name
        if match_type == "exact" and unit_name == pattern:
            return big_org, unit_name
        if match_type == "contains" and pattern in unit_name:
            return big_org, unit_name
    return unit_name, unit_name


def merge_bank_files_advanced(bank_dir, output_dir,
                              merge_different_banks=True,
                              merge_different_months=True,
                              merge_by_big_org=True,
                              filename_simple=True,
                              progress_callback=None):
    """
    高级合并银行报盘文件，支持三种合并策略开关。

    参数:
        bank_dir: 银行报盘目录（包含 YYYYMM-银行-单位名.xls 文件）
        output_dir: 输出目录
        merge_different_banks: True=同单位不同银行合并到同一文件
        merge_different_months: True=同单位不同月份合并到同一文件
        merge_by_big_org: True=按大单位合并
        progress_callback: (current, total, message)

    返回:
        (output_files, warnings, stats)
        output_files: 生成的合并文件路径列表
        warnings: 警告列表
        stats: 统计信息 dict
    """
    import os, re
    from collections import defaultdict

    bank_dir = os.path.normpath(bank_dir)
    output_dir = os.path.normpath(output_dir)
    os.makedirs(output_dir, exist_ok=True)

    bank_files = [f for f in os.listdir(bank_dir) if f.lower().endswith(".xls")]
    if not bank_files:
        return [], ["目录中未找到 .xls 文件"], {}

        if progress_callback:
            opts = []
            opts.append("同单位不同银行合并" if merge_different_banks else "同单位不同银行拆开")
            opts.append("同单位不同月份合并" if merge_different_months else "同单位不同月份拆开")
            opts.append("按大单位合并" if merge_by_big_org else "不按大单位合并")
            opts.append("简洁文件名" if filename_simple else "详细文件名")
            progress_callback(0, 1, f"扫描到 {len(bank_files)} 个报盘文件，规则：{'、'.join(opts)}")

    file_records = []
    skip_files = []
    for fname in sorted(bank_files):
        try:
            yearmon, bank, unit_name = split_filename(fname)
            file_records.append((yearmon, bank, unit_name, fname))
        except ValueError:
            skip_files.append(fname)

    if not file_records:
        return [], ["所有文件均无法解析文件名"], {}

    classified = []
    for yearmon, bank, unit_name, fname in file_records:
        big_org, _ = get_big_org(unit_name)
        excluded = is_excluded(unit_name)
        classified.append((yearmon, bank, unit_name, fname, big_org, excluded))

    groups = defaultdict(list)
    for rec in classified:
        yearmon, bank, unit_name, fname, big_org, excluded = rec
        if merge_by_big_org and not excluded:
            group_key = big_org
        else:
            group_key = unit_name
        groups[group_key].append(rec)

    _script_dir = os.path.dirname(os.path.abspath(__file__))
    bank_tmpl_path = os.path.join(_script_dir, "template", "代发业务导入模板.xlsx")
    if not os.path.exists(bank_tmpl_path):
        return [], [f"模板文件不存在: {bank_tmpl_path}"], {}

    warnings_list = []
    output_files = []
    all_operation_records = []  # 收集每条数据的来龙去脉
    stats = {
        "total_files": len(file_records),
        "total_groups": len(groups),
        "excluded_count": sum(1 for _, _, _, _, _, ex in classified if ex),
        "excluded_units": set(u for _, _, u, _, _, ex in classified if ex),
        "big_orgs": set(),
        "bank_counts": defaultdict(int),
    }

    total_groups = len(groups)
    pad_width = len(str(total_groups))

    # 银行排序
    def _bank_sort(b):
        order = {"建设银行": 0, "工商银行": 1, "吉林银行": 2,
                 "中国银行": 3, "交通银行": 4, "农业银行": 5,
                 "九台农商行": 6, "春城农商行": 7}
        return order.get(b, 99)

    for group_idx, (group_key, recs) in enumerate(sorted(groups.items()), 1):
        all_banks = sorted(set(r[1] for r in recs), key=_bank_sort)
        all_yearmons = sorted(set(r[0] for r in recs))
        if progress_callback:
            progress_callback(group_idx, total_groups,
                              f"合并: {group_key}（{len(recs)}个文件，{len(all_yearmons)}个月份，{len(all_banks)}家银行）")

        sub_groups = {}
        for rec in recs:
            yearmon, bank = rec[0], rec[1]
            if not merge_different_months and not merge_different_banks:
                sub_key = f"{yearmon}|{bank}"
            elif not merge_different_months:
                sub_key = yearmon
            elif not merge_different_banks:
                sub_key = bank
            else:
                sub_key = "_all"
            sub_groups.setdefault(sub_key, []).append(rec)

        def _sub_sort_key(k):
            if k == "_all":
                return (0, "")
            parts = k.split("|")
            if len(parts) == 2:
                return (1, parts[0], _bank_sort(parts[1]))
            if k in all_banks:
                return (2, _bank_sort(k))
            return (3, k)

        sorted_sub_keys = sorted(sub_groups.keys(), key=_sub_sort_key)

        for sub_key in sorted_sub_keys:
            sub_recs = sub_groups[sub_key]

            # 读取所有数据
            all_rows = []
            bank_counts = defaultdict(int)
            yearmons_in_sub = set()

            records_before = len(all_operation_records)  # 标记本轮起点

            for rec in sub_recs:
                yearmon, bank, unit_name, fname = rec[0], rec[1], rec[2], rec[3]
                rec_big_org = rec[4]
                rec_excluded = rec[5]
                fpath = os.path.join(bank_dir, fname)
                bt = detect_bank_type(fname)
                if bt == "icbc":
                    rows = _read_icbc_rows(fpath, warnings_list)
                elif bt == "ccb":
                    rows = _read_ccb_rows(fpath, warnings_list)
                elif bt == "jlb":
                    rows = _read_jlb_rows(fpath, warnings_list)
                else:
                    continue
                for row_data in rows:
                    op_rec = {
                        "source_file": fname,
                        "source_unit": unit_name,
                        "source_bank": bank,
                        "source_yearmon": yearmon,
                        "source_seq": row_data[0],
                        "name": str(row_data[1]).strip() if row_data[1] else "",
                        "id_number": str(row_data[2]).strip() if row_data[2] else "",
                        "amount": row_data[3],
                        "big_org": group_key,
                        "is_excluded": rec_excluded,
                        "output_file": "",
                        "output_seq": 0,
                        "filtered_reason": "",
                    }
                    is_zero = row_data[3] is not None and float(row_data[3]) == 0
                    if is_zero:
                        op_rec["filtered_reason"] = "金额为0，已过滤"
                    else:
                        all_rows.append(row_data)
                    all_operation_records.append(op_rec)
                bank_counts[bank] += len(rows)
                yearmons_in_sub.add(yearmon)

            # 同单位同月份同人同样金额重复检测
            dup_groups = defaultdict(list)
            for rec in all_operation_records[records_before:]:
                if rec.get("filtered_reason"):
                    continue
                person = str(rec.get("id_number", "")).strip()
                account = str(rec.get("name", "")).strip()
                if not person and not account:
                    continue
                key = (rec["source_yearmon"], rec["source_unit"],
                       person or "", account or "", float(rec["amount"]))
                dup_groups[key].append(rec)
            dup_warnings = {k: v for k, v in dup_groups.items() if len(v) > 1}
            if dup_warnings:
                for (ym, unit, person, account, amount), recs in dup_warnings.items():
                    source_files = sorted(set(r["source_file"] for r in recs))
                    acct_display = f" 账号{account}" if account else ""
                    warnings_list.append(
                        f"可能重复报盘：{unit} 月份{ym} "
                        f"户名「{person}」{acct_display}金额{float(amount):.2f}元 "
                        f"在{len(recs)}条记录中出现"
                        f"（来源：{'、'.join(source_files)}），请确认是否重复报盘"
                    )

            if not all_rows:
                continue

            for i, row in enumerate(all_rows, 1):
                row[0] = i

            sorted_ym = sorted(yearmons_in_sub)
            if len(sorted_ym) == 1:
                ym_suffix = sorted_ym[0]
            elif len(sorted_ym) > 1:
                try:
                    nums = sorted(int(y) for y in sorted_ym)
                    if all(nums[i] + 1 == nums[i+1] for i in range(len(nums)-1)):
                        ym_suffix = f"{sorted_ym[0]}-{sorted_ym[-1]}"
                    else:
                        ym_suffix = "_".join(sorted_ym)
                except ValueError:
                    ym_suffix = "_".join(sorted_ym)
            else:
                ym_suffix = ""

            seq_prefix = str(group_idx).zfill(pad_width)
            banks_in_sub = sorted(set(r[1] for r in sub_recs), key=_bank_sort)

            if filename_simple:
                parts = [group_key, ym_suffix] if ym_suffix else [group_key]
                merged_name = "-".join(parts) + ".xlsx"
            else:
                bank_parts = []
                for b in banks_in_sub:
                    count = bank_counts.get(b, 0)
                    bank_parts.append(f"{b}{count}笔")
                bank_suffix = "_".join(bank_parts)

                total_fen = 0
                for row in all_rows:
                    amt = row[3]
                    if isinstance(amt, Decimal):
                        total_fen += int(amt * 100)
                    else:
                        total_fen += int(round(float(amt) * 100))
                yuan = total_fen // 100
                jiao = (total_fen % 100) // 10
                fen = total_fen % 10

                parts = [f"{seq_prefix}_{group_key}", ym_suffix, bank_suffix,
                         f"总数{len(all_rows)}笔", f"总计{yuan}元{jiao}角{fen}分"]
                merged_name = "-".join(parts) + ".xlsx"
            merged_path = os.path.join(output_dir, merged_name)

            wb = openpyxl.load_workbook(bank_tmpl_path)
            ws = wb["代发工资模板"]
            for r in range(5, 8):
                for c in range(1, 7):
                    ws.cell(row=r, column=c).value = None
            for out_idx, row in enumerate(all_rows, start=5):
                ws.cell(row=out_idx, column=1, value=out_idx - 4)
                ws.cell(row=out_idx, column=2, value=str(row[1]).strip())
                ws.cell(row=out_idx, column=3, value=str(row[2]).strip())
                cell_e = ws.cell(row=out_idx, column=5, value=row[3])
                cell_e.number_format = "0.00"
            total_amount = sum(float(row[3] or 0) for row in all_rows)
            ws["B2"] = total_amount
            ws["B2"].number_format = "0.00"
            ws["B3"] = len(all_rows)
            wb.active = wb.sheetnames.index("代发工资模板")
            wb.save(merged_path)
            output_files.append(merged_path)

            written_idx = 0
            for rec in all_operation_records[records_before:]:
                if rec.get("filtered_reason"):
                    continue
                written_idx += 1
                rec["output_seq"] = written_idx
                rec["output_file"] = merged_name

            stats["big_orgs"].add(group_key)
            for b in banks_in_sub:
                stats["bank_counts"][b] += 1

    stats["output_count"] = len(output_files)
    stats["big_orgs"] = sorted(stats["big_orgs"])

    if skip_files:
        warnings_list.append(f"以下 {len(skip_files)} 个文件文件名不符合格式，已跳过：{', '.join(skip_files[:5])}")

    # 生成操作记录 Excel
    if all_operation_records:
        op_record_path = _generate_operation_record(output_dir, all_operation_records, stats, output_files)
        if op_record_path:
            warnings_list.append(f"操作记录已生成: {os.path.basename(op_record_path)}")

    if progress_callback:
        progress_callback(total_groups, total_groups, f"完成！生成 {len(output_files)} 个合并文件")

    return output_files, warnings_list, stats


def _generate_operation_record(output_dir, records, stats, output_files):
    """生成操作记录 Excel，追踪每条数据从哪个原始文件到哪个输出文件。"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

    if not records:
        return None

    out_path = os.path.join(output_dir, "操作记录.xlsx")
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "合并映射明细"

    thin = Side(style='thin')
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_align = Alignment(vertical='center', wrap_text=False)

    headers = [
        "序号", "源文件名", "源单位名", "源银行", "源月份",
        "源文件行号", "收款户名", "收款账号", "金额",
        "归类大单位", "是否排除",
        "输出文件名", "输出文件行号", "备注",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    col_widths = [8, 52, 36, 12, 10, 12, 14, 22, 12, 24, 10, 52, 14, 18]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + c) if c <= 26 else "?"].width = w

    for i, rec in enumerate(records, 1):
        r = i + 1
        ws.cell(r, 1, i).alignment = data_align
        ws.cell(r, 2, rec["source_file"]).alignment = data_align
        ws.cell(r, 3, rec["source_unit"]).alignment = data_align
        ws.cell(r, 4, rec["source_bank"]).alignment = data_align
        ws.cell(r, 5, rec["source_yearmon"]).alignment = data_align
        ws.cell(r, 6, rec["source_seq"]).alignment = data_align
        ws.cell(r, 7, rec["name"]).alignment = data_align
        ws.cell(r, 8, rec["id_number"]).alignment = data_align
        amt_cell = ws.cell(r, 9, float(rec["amount"]) if rec["amount"] else 0)
        amt_cell.number_format = "0.00"
        amt_cell.alignment = data_align
        ws.cell(r, 10, rec["big_org"]).alignment = data_align
        ws.cell(r, 11, "是" if rec["is_excluded"] else "").alignment = data_align
        ws.cell(r, 12, rec["output_file"]).alignment = data_align
        ws.cell(r, 13, rec["output_seq"]).alignment = data_align
        filtered = rec.get("filtered_reason", "")
        ws.cell(r, 14, filtered).alignment = data_align
        for c in range(1, 15):
            ws.cell(r, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
            if filtered:
                ws.cell(r, c).fill = PatternFill("solid", fgColor="FCE4EC")
            elif rec["is_excluded"]:
                ws.cell(r, c).fill = PatternFill("solid", fgColor="F2F2F2")

    ws.auto_filter.ref = f"A1:N{len(records) + 1}"
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("汇总")
    summary_rows = [
        ("统计项", "值"),
        ("源文件总数", stats.get("total_files", 0)),
        ("数据总行数", len(records)),
        ("金额为0已过滤行数", sum(1 for r in records if r.get("filtered_reason"))),
        ("排除项数据行数", sum(1 for r in records if r["is_excluded"])),
        ("生成合并文件数", len(output_files)),
        ("", ""),
        ("输出文件清单", ""),
    ]
    for path in sorted(output_files):
        basename = os.path.basename(path)
        group_rows = sum(1 for r in records if r["output_file"] == basename)
        summary_rows.append((basename, f"{group_rows} 行"))

    for r, (label, val) in enumerate(summary_rows, 1):
        c1 = ws2.cell(r, 1, label)
        c2 = ws2.cell(r, 2, val)
        if r == 1:
            c1.font = header_font
            c1.fill = header_fill
            c2.font = header_font
            c2.fill = header_fill
        if label == "":
            c1.font = Font(size=8)
            c2.font = Font(size=8)
    ws2.column_dimensions['A'].width = 60
    ws2.column_dimensions['B'].width = 20

    wb.save(out_path)
    return out_path


def convert_bank_format(bank_dir, output_dir):
    """
    读取 bank_dir 下所有银行报盘文件（工行/建行/吉林银行），
    合并为一个报盘，按吉林银行模板格式（代发业务导入模板.xlsx）输出。
    返回: (output_path, total_count) 或 (None, 0)
    """
    bank_dir = os.path.normpath(bank_dir)
    output_dir = os.path.normpath(output_dir)
    os.makedirs(output_dir, exist_ok=True)

    bank_files = [f for f in os.listdir(bank_dir) if f.lower().endswith(".xls")]
    if not bank_files:
        return None, 0

    all_rows = []
    warnings = []
    for fname in sorted(bank_files):
        fpath = os.path.join(bank_dir, fname)
        bt = detect_bank_type(fname)
        if bt == "icbc":
            rows = _read_icbc_rows(fpath, warnings)
        elif bt == "ccb":
            rows = _read_ccb_rows(fpath, warnings)
        elif bt == "jlb":
            rows = _read_jlb_rows(fpath, warnings)
        else:
            continue
        all_rows.extend(rows)

    if not all_rows:
        return None, 0

    for i, row in enumerate(all_rows, 1):
        row[0] = i

    _script_dir = os.path.dirname(os.path.abspath(__file__))
    tmpl_path = os.path.join(_script_dir, "template", "代发业务导入模板.xlsx")
    if not os.path.exists(tmpl_path):
        return None, 0

    # 从文件名提取年月做输出文件名
    yearmon = ""
    for fname in bank_files:
        try:
            y, _, _ = split_filename(fname)
            yearmon = y
            break
        except ValueError:
            continue
    if not yearmon:
        from datetime import datetime
        yearmon = datetime.now().strftime("%Y%m")

    out_name = f"合并报盘_{yearmon}.xlsx"
    out_path = os.path.normpath(os.path.join(output_dir, out_name))

    tmpl = openpyxl.load_workbook(tmpl_path)
    ws = tmpl["代发工资模板"]

    # 清除示例数据行
    for r in range(5, 8):
        for c in range(1, 7):
            ws.cell(row=r, column=c).value = None

    for out_idx, row in enumerate(all_rows, start=5):
        ws.cell(row=out_idx, column=1, value=out_idx - 4)
        ws.cell(row=out_idx, column=2, value=str(row[1]).strip())
        ws.cell(row=out_idx, column=3, value=str(row[2]).strip())
        cell_e = ws.cell(row=out_idx, column=5, value=row[3])
        cell_e.number_format = "0.00"

    last_data_row = 4 + len(all_rows)
    ws["B2"] = f"=SUM(E5:E{last_data_row})"
    ws["B3"] = f"=COUNT(A5:A{last_data_row})"
    tmpl.save(out_path)

    return out_path, len(all_rows)


def match_payroll_files(merged_files_list, payroll_dir):
    """
    输入：合并后的银行报盘文件列表（已排序）和工资表目录
    返回：(matched, unmatched, duplicates, no_signed_list)
    matched: [(merged_filename, payroll_filepath, unit_name), ...]
    unmatched: [payroll_filename, ...] — 有工资表但无对应银行报盘的单位
    duplicates: {unit_name: [filenames, ...], ...}
    no_signed_list: [(unit_name, unsigned_fname), ...] — 有银行报盘但工资表未签字的单位
    """
    result = []

    # 扫描工资表目录，建立索引
    payroll_files = []
    for fname in os.listdir(payroll_dir):
        # 排除汇总表和验证文件
        if "汇总表" in fname or "验证" in fname:
            continue
        payroll_files.append(fname)

    # 按单位名建立索引：unit_name -> [(is_signed, filename), ...]
    # 先按 signed_ 分组，再按扩展名优先级排序
    from collections import defaultdict
    payroll_index = defaultdict(list)

    for fname in payroll_files:
        unit_candidate = None
        is_signed = fname.startswith("signed_")

        # 去掉 signed_ 前缀用于匹配
        base = fname[len("signed_"):] if is_signed else fname

        # 从工资表文件名中提取单位名（去掉年月后缀部分）
        # 格式: 吉林大学数学学院2026年06月工资表.xlsx → 吉林大学数学学院
        # 或: (吉林大学）深部探测与成像全国重点实验室202606工资.xls → (吉林大学）深部探测与成像全国重点实验室
        for sep in ["2026年06月工资表", "202606工资表", "202606工资"]:
            if sep in base:
                unit_candidate = base.split(sep, 1)[0].strip()
                break

        if unit_candidate:
            payroll_index[unit_candidate].append((is_signed, fname))

    # 对每个合并文件进行匹配
    matched_units = set()
    duplicates = {}  # unit_name -> [all matched filenames]
    no_signed_list = []  # 有银行报盘但工资表未签字的单位
    for merged_name in merged_files_list:
        # 提取单位名：第一个 '-202606-' 之前的部分（去掉序号前缀）
        raw_unit = merged_name.split("-202606-", 1)[0]
        # 去掉开头的序号前缀 "001_"
        unit_name = raw_unit.split("_", 1)[1] if "_" in raw_unit else raw_unit

        matched_file = None

        if unit_name in payroll_index:
            matched_units.add(unit_name)
            candidates = payroll_index[unit_name]
            # 记录所有候选文件，用于重复检测
            all_candidates = [f for _, f in candidates]
            if len(all_candidates) > 1:
                duplicates[unit_name] = all_candidates

            # 优先级：signed_ > 非 signed_，同优先级下 .xlsx > .xls
            signed = [f for s, f in candidates if s]
            unsigned = [f for s, f in candidates if not s]

            if signed:
                # signed_ 中优先选 .xlsx
                xlsx = [f for f in signed if f.endswith(".xlsx")]
                matched_file = xlsx[0] if xlsx else signed[0]
            elif unsigned:
                # 非 signed_ 不视为正式工资表，记录到 no_signed_list
                no_signed_list.append((unit_name, unsigned[0]))

        filepath = os.path.join(payroll_dir, matched_file) if matched_file else None
        result.append((merged_name, filepath, unit_name))

    # 反向检查：工资表中有哪些单位没被匹配到
    unmatched = []
    for unit_name, candidates in payroll_index.items():
        if unit_name not in matched_units:
            # 取第一个文件名作为代表
            _, fname = candidates[0]
            unmatched.append(fname)

    return result, unmatched, duplicates, no_signed_list


# ──────────────────────────────────────────────
# WPS 打印模块
# ──────────────────────────────────────────────


def check_wps_available():
    """
    检测 WPS Office COM 组件是否可用
    返回: True/False
    """
    try:
        import win32com.client
        app = win32com.client.Dispatch("KET.Application")
        app.Quit()
        return True
    except (ImportError, Exception):
        return False


def print_file(filepath, progress_callback=None):
    """
    打印单个 Excel 文件
    参数:
      filepath: 文件路径
      progress_callback: 可选回调(c, t, msg)，用于 GUI 进度更新
    返回: True/False
    """
    import time

    try:
        import win32com.client
    except ImportError:
        if progress_callback:
            progress_callback(0, 1, "win32com 不可用（非 Windows 环境）")
        return False

    max_retries = 3
    for attempt in range(1, max_retries + 1):
        app = None
        try:
            app = win32com.client.DispatchEx("KET.Application")
            app.Visible = False
            app.DisplayAlerts = False  # 禁止所有弹窗

            wb = app.Workbooks.Open(filepath)
            ws = wb.ActiveSheet

            # WPS 页面设置：A4 横向 + 所有列缩放到一页 + 每页重复表头
            ws.PageSetup.Orientation = 2          # xlLandscape
            ws.PageSetup.PaperSize = 9            # xlPaperA4
            # Zoom 不设值，FitToPagesWide 已自动禁用百分比缩放
            ws.PageSetup.FitToPagesWide = 1       # 所有列缩放到一页宽
            ws.PageSetup.FitToPagesTall = 0       # 0=不限制行高页数
            ws.PageSetup.PrintTitleRows = "$1:$5" # 每页重复表头

            # 让含图片的行按内容自动调整行高，避免 PDF 行高压缩
            try:
                for _shape in ws.Shapes:
                    if _shape.Type == 13:  # msoPicture
                        _row = _shape.TopLeftCell.Row
                        ws.Rows(_row).RowHeight = 50
            except Exception:
                pass

            ws.PrintOut()

            wb.Close(SaveChanges=False)

            return True

        except Exception as e:
            import traceback as _tb
            err_detail = f"{type(e).__name__}: {e}"
            if progress_callback:
                progress_callback(0, 1, f"✗ 打印失败 (第{attempt}次): {err_detail}")
                progress_callback(0, 1, f"  traceback: {_tb.format_exc()[:500]}")
            if attempt < max_retries:
                time.sleep(5)
            else:
                return False
        finally:
            if app is not None:
                try:
                    app.Quit()
                except Exception:
                    pass
            time.sleep(2.5)


def batch_print(matched_pairs, progress_callback=None):
    """
    批量打印工资表
    参数:
      matched_pairs: [(merged_filename, payroll_filepath, unit_name), ...]
      progress_callback: 可选回调函数(current, total, message)
    返回: (success_count, fail_count, fail_list)
    """
    if not check_wps_available():
        return (0, 0, ["WPS不可用"])

    total = len(matched_pairs)
    success_count = 0
    fail_count = 0
    fail_list = []

    for i, (merged_filename, payroll_filepath, unit_name) in enumerate(matched_pairs):
        if progress_callback:
            progress_callback(i + 1, total, f"正在打印: {unit_name}")

        ok = print_file(payroll_filepath, progress_callback)
        if ok:
            success_count += 1
        else:
            fail_count += 1
            fail_list.append((merged_filename, unit_name))

    return (success_count, fail_count, fail_list)


def generate_report_xlsx(output_dir, renamed, matched, unmatched, duplicates,
                         merge_warnings=None, success=None, fail=None, fail_list=None,
                         no_signed=None):
    """
    生成合并打印操作记录 xlsx
    renamed: [(new_name, old_name, bank_name, unit_name), ...]
    matched: [(merged_filename, payroll_filepath, unit_name), ...]
    unmatched: [payroll_filename, ...]
    duplicates: {unit_name: [filenames, ...], ...}
    merge_warnings: [str, ...] 金额警告
    success/fail/fail_list: 打印结果（仅打印模式有值）
    返回: 报告文件路径
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = os.path.join(output_dir, f"操作记录_{ts}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "操作记录"

    # ── 标题行 ──
    ws.cell(row=1, column=1, value="合并打印操作记录")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=14)

    # ── 汇总信息 ──
    row = 3
    ws.cell(row=row, column=1, value="银行报盘文件数")
    ws.cell(row=row, column=2, value=len(renamed))
    row += 1
    ws.cell(row=row, column=1, value="合并文件数")
    ws.cell(row=row, column=2, value=len(matched))
    row += 1
    if success is not None:
        ws.cell(row=row, column=1, value="打印成功数")
        ws.cell(row=row, column=2, value=success)
        row += 1
        ws.cell(row=row, column=1, value="打印失败数")
        ws.cell(row=row, column=2, value=fail)
        row += 1
    ws.cell(row=row, column=1, value="未匹配到报盘的工资表数")
    ws.cell(row=row, column=2, value=len(unmatched))
    row += 1
    ws.cell(row=row, column=1, value="存在多个工资表的单位数")
    ws.cell(row=row, column=2, value=len(duplicates))
    row += 1
    if no_signed:
        ws.cell(row=row, column=1, value="工资表未签字（无 signed_ 前缀）的单位数")
        ws.cell(row=row, column=2, value=len(no_signed))
        row += 1
        ws.cell(row=row, column=1, value="未签字明细")
        ws.cell(row=row, column=2, value="单位名称")
        ws.cell(row=row, column=3, value="工资表文件")
        for uname, fn in no_signed:
            row += 1
            ws.cell(row=row, column=2, value=uname)
            ws.cell(row=row, column=3, value=fn)

    # 建立原始报盘文件查找表：unit_name -> [old_name, ...]
    orig_files = {}
    for new_name, old_name, bank_name, unit_name in renamed:
        orig_files.setdefault(unit_name, []).append(old_name)

    # ── 明细表头 ──
    row += 2
    headers = ["序号", "单位名称", "合并文件", "原始报盘文件", "工资表文件", "匹配状态", "打印状态", "备注"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # 建立打印结果查找表：unit_name -> "成功"/"失败"/"未打印"
    print_status = {}
    if success is not None:
        # 默认未打印
        for _, _, unit_name in matched:
            print_status[unit_name] = "未打印"
        # 覆盖成功/失败
        for merged_name, payroll_path, unit_name in matched:
            if payroll_path is None:
                print_status[unit_name] = "未打印（无工资表）"
        # 如果有 fail_list，标记失败
        for merged_name, unit_name in (fail_list or []):
            print_status[unit_name] = "失败"
        # 其余标记成功（排除了无工资表的和失败的）
        for merged_name, payroll_path, unit_name in matched:
            if payroll_path and print_status.get(unit_name) == "未打印":
                print_status[unit_name] = "成功"
    else:
        for _, _, unit_name in matched:
            print_status[unit_name] = "未打印"

    # ── 明细数据 ──
    no_signed_units = set(u for u, _ in (no_signed or []))
    for idx, (merged_name, payroll_path, unit_name) in enumerate(matched, 1):
        row += 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=unit_name)
        ws.cell(row=row, column=3, value=merged_name)
        # 原始报盘文件
        orig_list = orig_files.get(unit_name, [])
        ws.cell(row=row, column=4, value="\n".join(orig_list))
        ws.cell(row=row, column=5, value=os.path.basename(payroll_path) if payroll_path else "")
        if payroll_path:
            status_text = "已匹配"
        elif unit_name in no_signed_units:
            status_text = "未签字"
        else:
            status_text = "未找到"
        ws.cell(row=row, column=6, value=status_text)
        ws.cell(row=row, column=7, value=print_status.get(unit_name, ""))
        # 备注：重复文件提示
        note = ""
        if unit_name in duplicates:
            note = f"存在多个工资表文件，仅使用 {os.path.basename(payroll_path) if payroll_path else '第一个'}"
        ws.cell(row=row, column=8, value=note)

    # ── 金额警告 ──
    if merge_warnings:
        row += 2
        ws.cell(row=row, column=1, value="金额警告（分以下数值已舍去）：")
        ws.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True, color="FF8C00")
        for w in merge_warnings:
            row += 1
            ws.cell(row=row, column=1, value=w)

    # ── 未匹配工资表 ──
    if unmatched:
        row += 2
        ws.cell(row=row, column=1, value="未匹配到银行报盘的工资表：")
        ws.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True, color="FF0000")
        for fname in unmatched:
            row += 1
            ws.cell(row=row, column=1, value=fname)

    # 列宽
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 45

    wb.save(report_path)
    return report_path


# ──────────────────────────────────────────────
# 工资表合并（按个税分组）
# ──────────────────────────────────────────────


def _read_payroll_data(filepath):
    """读取工资表文件，返回 (header_rows, data_rows, footer_rows, tax_col_idx)
    header_rows: 表头行列表（每行是单元格值列表）
    data_rows: 数据行列表（每行是单元格值列表）
    footer_rows: 表尾行列表（备注行、签字行等）
    tax_col_idx: 个人所得税列索引（0-based）
    """
    is_xls = filepath.lower().endswith(".xls")
    if is_xls:
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        nrows, ncols = ws.nrows, ws.ncols
    else:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        nrows, ncols = ws.max_row, ws.max_column

    # 找个人所得税列
    tax_col = None
    for c in range(ncols):
        if is_xls:
            val = str(ws.cell_value(2, c)).strip()
        else:
            val = str(ws.cell(row=3, column=c + 1).value or "").strip()
        if "个人所得" in val or "个税" in val:
            tax_col = c
            break

    # 读取所有行
    all_rows = []
    for r in range(nrows):
        row = []
        for c in range(ncols):
            if is_xls:
                row.append(ws.cell_value(r, c))
            else:
                row.append(ws.cell(row=r + 1, column=c + 1).value)
        all_rows.append(row)

    # 分离表头、数据、表尾
    # 表头：行0-4（标题行、单位行、列名行、扣款子行1、扣款子行2）
    # 数据行：有序号（第1列是数字）且姓名列有值且不是"合计"的行
    # 表尾：其余行（备注行、签字行、空行）
    header_rows = all_rows[:5]
    data_rows = []
    footer_rows = []
    for row in all_rows[5:]:
        seq = str(row[0]).strip() if len(row) > 0 else ""
        name = str(row[1]).strip() if len(row) > 1 else ""
        # 数据行特征：序号是数字，姓名有值且不是合计/转账等汇总行
        is_data = False
        try:
            float(seq)  # 序号是数字
            if name and name not in ("合计", "转账合计"):
                is_data = True
        except (ValueError, TypeError):
            pass
        if is_data:
            data_rows.append(row)
        else:
            footer_rows.append(row)

    return header_rows, data_rows, footer_rows, tax_col


def _get_common_columns(all_headers):
    """从所有表头中找出共同列和独有列
    返回: (common_cols, unique_cols_map)
    common_cols: [(idx, name), ...] 所有文件都有的列
    unique_cols_map: {filename: [(idx, name), ...]} 每个文件独有的列
    """
    # 取第一个文件的列名作为基准
    base = all_headers[0][2]  # 第3行是列名行
    base_names = [str(v)[:20] for v in base]

    common = list(range(len(base_names)))
    unique_map = {}

    for i, headers in enumerate(all_headers[1:], 1):
        names = [str(v)[:20] for v in headers[2]]
        # 找当前文件比基准多的列
        extra = []
        for c, name in enumerate(names):
            if name and name not in base_names:
                extra.append((c, name))
        if extra:
            unique_map[i] = extra

    return [(i, base_names[i]) for i in range(len(base_names))], unique_map


def _detect_column_structure(headers):
    """检测工资表的列结构类型，返回列名列表和特殊列信息"""
    row3 = [str(v or "").strip() for v in headers[2]]
    row4 = [str(v or "").strip() for v in headers[3]] if len(headers) > 3 else []
    row5 = [str(v or "").strip() for v in headers[4]] if len(headers) > 4 else []

    # 构建列名（合并行3-5的信息）
    col_names = []
    for c in range(len(row3)):
        name = row3[c]
        # 扣款明细主列名
        if "扣" in name and "款" in name:
            name = "扣款明细"
        col_names.append(name)

    # 检测特殊列
    special_cols = {}
    for c, name in enumerate(col_names):
        if "补发工资" in name:
            special_cols["补发工资"] = c
        elif "大病险" in name:
            special_cols["大病险"] = c
        elif "雇主责任险" in name:
            special_cols["雇主责任险"] = c
        elif "公务员医疗补助" in name:
            special_cols["公务员医疗补助"] = c
        elif "单位缴纳五险一金" in name:
            special_cols["单位缴纳五险一金"] = c

    # 检查行4/5中是否有公务员医疗补助
    for r in [row4, row5]:
        for c, v in enumerate(r):
            if "公务员医疗补助" in v:
                special_cols["公务员医疗补助"] = c

    return col_names, special_cols


def _build_universal_column_map(col_names, special_cols):
    """将任意列结构映射到统一输出列（删除部门/岗位/职工号，保留特殊列）
    返回: (output_cols, delete_indices)
    output_cols: 输出列索引列表（0-based），每个元素为 (输出列索引, 原始列索引)
    delete_indices: 被删除的原始列索引
    """
    # 标准列顺序（所有文件都有）
    # 序号(0), 姓名(1), 身份证(2), 部门(3), 岗位(4), 职工号(5),
    # 基本工资(6), 交通补贴(7) / 补发工资(7), 应发工资(8/9),
    # 单位缴纳五险一金(9/10), 单位代理费(10/11), 大病险(11/12),
    # 雇主责任险(11/12), 转账合计(11/12/13),
    # 社保基数(12/13/14), 医保基数(13/14/15), 工伤基数(14/15/16), 公积金基数(15/16/17),
    # 扣款明细(16-26/17-27), 个人所得税(27/28), 实发工资(28/29), 实发合计(29/30)

    # 删除列：部门(3), 岗位(4), 职工号(5)
    delete_indices = {3, 4, 5}

    # 构建输出映射
    output_cols = []
    for c in range(len(col_names)):
        if c in delete_indices:
            continue
        output_cols.append(c)

    return output_cols, delete_indices


def _read_payroll_workbook(filepath):
    """读取工资表文件，返回 (wb_or_ws, is_xls, nrows, ncols, merged_cells, images)"""
    is_xls = filepath.lower().endswith(".xls")
    if is_xls:
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        return (wb, ws, is_xls, ws.nrows, ws.ncols, [], [])
    else:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        merged = list(ws.merged_cells.ranges) if ws.merged_cells else []
        # 从 xlsx zip 中直接读取图片文件（openpyxl 的 img.path 在 3.1.5 中返回错误路径）
        images = []
        if ws._images:
            import zipfile
            from xml.etree import ElementTree as ET
            try:
                with zipfile.ZipFile(filepath) as z:
                    # 解析 drawing rels: rId → 真实图片路径
                    try:
                        rels_xml = z.read("xl/drawings/_rels/drawing1.xml.rels")
                        root = ET.fromstring(rels_xml)
                        ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
                        rid_map = {}
                        for rel in root.findall("r:Relationship", ns):
                            rid = rel.get("Id")
                            target = rel.get("Target", "")
                            if rid and target:
                                rid_map[rid] = target.lstrip("/")
                    except Exception:
                        rid_map = {}

                    for img in ws._images:
                        try:
                            embed = img.anchor.pic.blipFill.blip.embed
                            png_path = rid_map.get(embed)
                            if not png_path:
                                png_path = img.path.lstrip("/")  # fallback
                            png_data = z.read(png_path)
                            images.append((png_data, img.anchor._from.row, img.anchor._from.col, img.width, img.height))
                        except Exception:
                            pass
            except Exception:
                pass
        return (wb, ws, is_xls, ws.max_row, ws.max_column, merged, images)


def _build_column_name_map(headers):
    """从表头构建列名→索引映射。
    Merged cells in row3/row4 produce empty subsequent entries; we track
    the last non-empty row3 and row4 values to create unique keys.
    返回 (name_to_idx, col_names):
    name_to_idx: {规范列名: 0-based索引}
    col_names: [规范列名, ...] 按原始列顺序
    """
    row3 = [str(v or "").strip() for v in headers[2]]
    row4 = [str(v or "").strip() for v in headers[3]] if len(headers) > 3 else []
    row5 = [str(v or "").strip() for v in headers[4]] if len(headers) > 4 else []

    name_to_idx = {}
    col_names = []
    last_row3 = ""
    last_row4 = ""

    for c in range(len(row3)):
        raw = row3[c]
        if raw and "扣" in raw and "款" in raw:
            name = "扣款明细"
        elif raw:
            name = raw
        else:
            name = last_row3
        if raw:
            last_row3 = name

        # Propagate row4 only within merged ranges (row3 empty)
        r4_val = ""
        if c < len(row4) and row4[c]:
            r4_val = row4[c]
            last_row4 = row4[c]
        elif c < len(row4) and not raw and last_row4:
            r4_val = last_row4
        elif c < len(row4) and raw:
            last_row4 = ""  # Reset when entering a new section

        sub_parts = [r4_val] if r4_val else []
        if c < len(row5) and row5[c]:
            sub_parts.append(row5[c])
        sub_name = "/".join(sub_parts) if sub_parts else ""

        if sub_name:
            key = f"{name}/{sub_name}" if name != sub_name else name
        else:
            key = name
        name_to_idx[key] = c
        col_names.append(key)

    return name_to_idx, col_names


def _get_canonical_columns(all_file_columns):
    """扫描所有文件，按列在各文件中的典型位置自动排顺序。
    返回: [规范列名, ...] 不包含 部门/岗位/职工号
    """
    # 统计每个列名在所有文件中出现的位置
    pos_map = {}
    for cols in all_file_columns:
        for i, c in enumerate(cols):
            if c in ("部门", "岗位", "职工号"):
                continue
            if c not in pos_map:
                pos_map[c] = []
            pos_map[c].append(i)

    # 用中位数做主排序，均值做辅排序（处理同中位数的列，如 补发工资 vs 交通补贴）
    def _median_low(values):
        s = sorted(values)
        return s[(len(s) - 1) // 2]

    col_order = {}
    for c, ps in pos_map.items():
        col_order[c] = (_median_low(ps), sum(ps) / len(ps))

    result = sorted(col_order.keys(), key=lambda c: col_order[c])
    return result


def _normalize_row_by_names(row, headers, canonical_cols):
    """根据规范列名将原始行映射为规范行。
    row: 原始数据行（值列表）
    headers: 该文件的表头
    canonical_cols: 规范列名列表
    返回: [值, ...] 按规范列顺序
    """
    name_to_idx, _ = _build_column_name_map(headers)
    result = []
    for cname in canonical_cols:
        if cname in name_to_idx:
            idx = name_to_idx[cname]
            if idx < len(row):
                result.append(row[idx])
            else:
                result.append("")
        else:
            result.append("")
    return result


def _adjust_merged_range(mrange, delete_1based_set, insert_before_1based=3):
    """调整合并单元格范围，处理列删除和插入。
    mrange: openpyxl merged cell range 或 str
    delete_1based_set: 被删除列的1-based列号集合
    insert_before_1based: 在1-based列号前插入新列（结算单元，位于姓名后=原C前）
    返回: 调整后的范围字符串，或 None
    """
    import re
    rng_str = str(mrange)
    parts = rng_str.split(":")
    if len(parts) != 2:
        return None

    from_cell, to_cell = parts[0], parts[1]
    m1 = re.match(r"([A-Z]+)(\d+)", from_cell)
    m2 = re.match(r"([A-Z]+)(\d+)", to_cell)
    if not m1 or not m2:
        return None

    from_col_str, from_row = m1.group(1), int(m1.group(2))
    to_col_str, to_row = m2.group(1), int(m2.group(2))

    from_col = sum((ord(c) - 64) * (26 ** i) for i, c in enumerate(reversed(from_col_str)))
    to_col = sum((ord(c) - 64) * (26 ** i) for i, c in enumerate(reversed(to_col_str)))

    def _mapped_col(col_1based):
        """Original 1-based → output 1-based, accounting for deletion then insertion."""
        # Step 1: deletion (3 cols removed at 1-based 4,5,6)
        deleted_before = sum(1 for d in sorted(delete_1based_set) if d < col_1based)
        after_delete = col_1based - deleted_before
        if after_delete <= 0:
            return None
        # Step 2: insertion (结算单元 inserted at 1-based column = insert_before_1based)
        if after_delete >= insert_before_1based:
            return after_delete + 1
        else:
            return after_delete

    new_from = _mapped_col(from_col)
    new_to = _mapped_col(to_col)

    if new_from is None or new_to is None or new_from > new_to:
        return None

    def col_to_letter(n):
        s = ""
        while n > 0:
            n, r = divmod(n - 1, 26)
            s = chr(65 + r) + s
        return s

    return f"{col_to_letter(new_from)}{from_row}:{col_to_letter(new_to)}{to_row}"


def _get_default_validation_config():
    """返回默认的验证规则配置（内置默认值，无配置文件时使用）。"""
    return {
        "tolerance": 0.005,
        "deduction": {
            "prefix": "扣款明细/",
            "exclude_keywords": ["扣款合计", "大病险合计"]
        },
        "row_formulas": [
            {
                "name": "转账合计 = 应发工资 + 单位缴纳五险一金 + 单位代理费 + 雇主责任险 + 大病险",
                "lhs": [["转账合计"]],
                "rhs": [
                    ["应发工资"],
                    ["单位缴纳五险一金"],
                    ["单位代理费"],
                    ["雇主责任险"],
                    ["大病险"]
                ],
                "check_per_row": true
            },
            {
                "name": "转账合计 = 扣款合计 + 个人所得税 + 实发合计",
                "lhs": [["转账合计"]],
                "rhs": [
                    ["扣款明细/扣款合计", "扣款合计"],
                    ["个人所得税"],
                    ["实发合计"]
                ],
                "check_per_row": true
            },
            {
                "name": "扣款合计 = 扣款明细子项之和",
                "lhs": [["扣款明细/扣款合计", "扣款合计"]],
                "rhs_subtract": true,
                "check_per_row": true
            }
        ],
        "column_sum_targets": [
            ["基本工资"], ["补发工资"], ["应发工资"],
            ["单位缴纳五险一金"], ["大病险"],
            ["转账合计"], ["个人所得税"], ["实发合计"]
        ],
        "critical_columns": {
            "转账合计": ["转账合计"],
            "扣款合计": ["扣款明细/扣款合计", "扣款合计"],
            "实发合计": ["实发合计"],
            "应发工资": ["应发工资"]
        }
    }


def _load_validation_config(config_file=None):
    """
    加载验证规则配置文件。若未指定或文件不存在，返回内置默认配置。
    """
    import json

    if config_file is None:
        config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "validation_config.json")

    if os.path.isfile(config_file):
        try:
            with open(config_file, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            return cfg
        except Exception as e:
            print(f"Warning: 无法读取验证配置文件 {config_file}，使用默认配置: {e}")

    return _get_default_validation_config()


def _save_validation_config(config, config_file=None):
    """保存验证规则配置文件。"""
    import json
    if config_file is None:
        config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "validation_config.json")
    with open(config_file, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)


def _load_summary_config():
    """加载结算单元汇总字段配置。若文件不存在，返回默认字段列表。"""
    import json
    config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "summary_config.json")
    if os.path.isfile(config_file):
        try:
            with open(config_file, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            if isinstance(cfg, dict) and "fields" in cfg:
                return cfg["fields"]
            if isinstance(cfg, list):
                return cfg
            return _get_default_summary_fields()
        except Exception as e:
            print(f"Warning: 无法读取汇总配置 {config_file}，使用默认: {e}")
    return _get_default_summary_fields()


def _save_summary_config(fields):
    """保存结算单元汇总字段配置。"""
    import json
    config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "summary_config.json")
    with open(config_file, "w", encoding="utf-8") as f:
        json.dump({"fields": fields}, f, ensure_ascii=False, indent=2)


def _get_default_summary_fields():
    """返回默认的结算单元汇总字段列表。"""
    return ["个人所得税", "个人工会会费", "工会经费", "实发合计", "实发工资"]


def validate_payroll_xlsx(filepath, canonical_cols, config=None, tolerance=None):
    """
    验证生成的工资表 xlsx 中合计行数据是否符合公式，并将结果追加为新 sheet。

    校验规则由 config 定义（dict 或配置文件路径），支持：
      - row_formulas: 行公式校验 (lhs = sum(rhs_plus) - sum(rhs_minus))
      - column_sum_targets: 列加总校验
      - critical_columns: 关键列存在性检查
      - deduction: 扣款子项汇总校验

    参数:
      filepath: xlsx 文件路径
      canonical_cols: 规范列名列表 (0-based 顺序)
      config: 配置 dict、配置文件路径、或 None（使用默认配置）
      tolerance: 比较容差，若指定则覆盖 config 中的值

    返回:
      {"ok": bool, "passed_count": int, "failed_count": int, "checks": [...]}
    """
    from openpyxl.styles import PatternFill, Font, Alignment

    # ── 加载配置 ──
    if config is None:
        cfg = _load_validation_config()
    elif isinstance(config, str):
        cfg = _load_validation_config(config)
    elif isinstance(config, dict):
        cfg = config
    else:
        cfg = _load_validation_config()

    tol = tolerance if tolerance is not None else cfg.get("tolerance", 0.005)
    deduction_cfg = cfg.get("deduction", {})
    row_formulas = cfg.get("row_formulas", [])
    col_sum_targets = cfg.get("column_sum_targets", [])
    critical_columns = cfg.get("critical_columns", {})

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    nrows = ws.max_row
    ncols = ws.max_column

    # ── 列索引 ──
    col_idx = {name: i for i, name in enumerate(canonical_cols)}

    def _col(keywords):
        for kw in keywords:
            if kw in col_idx:
                return col_idx[kw]
        return -1

    def _val(row_1based, col_0based):
        if col_0based < 0:
            return 0.0
        v = ws.cell(row=row_1based, column=col_0based + 1).value
        try:
            return round(float(v or 0), 2)
        except (ValueError, TypeError):
            return 0.0

    # ── 合计行 ──
    summary_row = None
    for r in range(1, nrows + 1):
        if str(ws.cell(row=r, column=1).value or "").strip() == "合计":
            summary_row = r
            break

    if summary_row is None:
        _append_validation_sheet(wb, [{
            "name": "合计行查找", "kind": "table_format",
            "passed": False, "detail": "未找到合计行"
        }])
        wb.save(filepath)
        wb.close()
        return {"ok": False, "passed_count": 0, "failed_count": 1,
                "checks": [{"name": "合计行查找", "passed": False}]}

    # ── 表头行数 ──
    last_header_row = 3
    for r in [4, 5]:
        if r <= nrows:
            has_content = any(
                str(ws.cell(row=r, column=c).value or "").strip()
                for c in range(1, ncols + 1) if c != 1
            )
            if has_content:
                last_header_row = r
            else:
                break

    data_start = last_header_row + 1
    data_end = summary_row - 1
    data_count = max(0, data_end - data_start + 1)

    # 姓名列索引（用于逐行报错时显示员工名）
    name_col_idx = _col(["姓名"])

    checks = []

    def _add_check(name, kind, passed, detail=""):
        checks.append({
            "name": name, "kind": kind,
            "passed": passed, "detail": detail,
        })

    def _row_name(r):
        """获取第 r 行员工姓名"""
        if name_col_idx >= 0:
            v = ws.cell(row=r, column=name_col_idx + 1).value
            return str(v or "").strip()
        return f"第{r}行"

    # ──────────────────────────────────────────
    # Row formula checks (行公式校验)
    # ──────────────────────────────────────────
    for formula in row_formulas:
        formula_name = formula.get("name", "公式校验")
        check_per_row = formula.get("check_per_row", False)

        if formula.get("rhs_subtract"):
            # 特殊规则：扣款合计 = 所有扣款明细子项之和
            ded_prefix = deduction_cfg.get("prefix", "扣款明细/")
            ded_exclude = deduction_cfg.get("exclude_keywords", ["扣款合计", "大病险合计"])

            lhs_keywords_list = formula.get("lhs", [["扣款明细/扣款合计", "扣款合计"]])
            lhs_idx = -1
            for kw_list in lhs_keywords_list:
                lhs_idx = _col(kw_list)
                if lhs_idx >= 0:
                    break

            sub_indices = [
                idx for name, idx in col_idx.items()
                if name.startswith(ded_prefix)
                and not any(excl in name for excl in ded_exclude)
            ]

            # ── 合计行检查 ──
            if sub_indices and lhs_idx >= 0:
                sub_sum = round(sum(_val(summary_row, idx) for idx in sub_indices), 2)
                lhs_val = _val(summary_row, lhs_idx)
                diff = round(sub_sum - lhs_val, 2)
                if abs(diff) <= tol:
                    _add_check(formula_name + " (合计行)", "row_formula", True)
                else:
                    _add_check(formula_name + " (合计行)", "row_formula", False,
                                f"子项和 {sub_sum:.2f} ≠ 合计 {lhs_val:.2f} (差 {diff:+.2f})")
            elif lhs_idx < 0:
                _add_check(formula_name + " (合计行)", "row_formula", True,
                            "（跳过：合计列不存在）")
            else:
                _add_check(formula_name + " (合计行)", "row_formula", True,
                            "（跳过：无扣款子项）")

            # ── 逐行检查 ──
            if check_per_row and sub_indices and lhs_idx >= 0:
                pass_count = 0
                fail_count = 0
                fail_samples = []
                for r in range(data_start, data_end + 1):
                    sub_sum_r = round(sum(_val(r, idx) for idx in sub_indices), 2)
                    lhs_val_r = _val(r, lhs_idx)
                    d = round(sub_sum_r - lhs_val_r, 2)
                    if abs(d) <= tol:
                        pass_count += 1
                    else:
                        fail_count += 1
                        if len(fail_samples) < 3:
                            fail_samples.append(f"{_row_name(r)}: 子项和{sub_sum_r:.2f}≠{lhs_val_r:.2f}")
                if fail_count == 0:
                    _add_check(formula_name + " (逐行)", "per_row", True,
                                f"{pass_count} 行全部通过")
                else:
                    sample_str = " | ".join(fail_samples)
                    _add_check(formula_name + " (逐行)", "per_row", False,
                                f"通过 {pass_count} 行，失败 {fail_count} 行 | {sample_str}")
        else:
            # 标准行公式：lhs = sum(rhs)
            lhs_idx = -1
            for kw_list in formula.get("lhs", []):
                lhs_idx = _col(kw_list)
                if lhs_idx >= 0:
                    break

            rhs_indices = []
            rhs_labels = []
            for kw_list in formula.get("rhs", []):
                idx = _col(kw_list)
                if idx >= 0:
                    rhs_indices.append(idx)
                    rhs_labels.append(kw_list[0])
                else:
                    rhs_indices.append(-1)
                    rhs_labels.append(kw_list[0])

            if lhs_idx < 0:
                all_rhs_missing = all(idx < 0 for idx in rhs_indices)
                if all_rhs_missing:
                    _add_check(formula_name + " (合计行)", "row_formula", True,
                                "（跳过：相关列均不存在）")
                else:
                    _add_check(formula_name + " (合计行)", "row_formula", False,
                                "LHS 列不存在")
                continue

            # ── 合计行检查 ──
            lhs_val = _val(summary_row, lhs_idx)
            rhs_sum = round(sum(_val(summary_row, idx) for idx in rhs_indices if idx >= 0), 2)
            diff = round(lhs_val - rhs_sum, 2)

            if abs(diff) <= tol:
                _add_check(formula_name + " (合计行)", "row_formula", True)
            else:
                rhs_parts = []
                for kw_list, idx in zip(formula.get("rhs", []), rhs_indices):
                    v = _val(summary_row, idx)
                    rhs_parts.append(f"{kw_list[0]}{v:.2f}")
                _add_check(formula_name + " (合计行)", "row_formula", False,
                            f"{lhs_val:.2f} ≠ {' + '.join(rhs_parts)} = {rhs_sum:.2f} (差 {diff:+.2f})")

            # ── 逐行检查 ──
            if check_per_row:
                pass_count = 0
                fail_count = 0
                fail_samples = []
                for r in range(data_start, data_end + 1):
                    lhs_val_r = _val(r, lhs_idx)
                    rhs_sum_r = round(sum(_val(r, idx) for idx in rhs_indices if idx >= 0), 2)
                    d = round(lhs_val_r - rhs_sum_r, 2)
                    if abs(d) <= tol:
                        pass_count += 1
                    else:
                        fail_count += 1
                        if len(fail_samples) < 3:
                            rhs_parts_r = []
                            for kw_list, idx in zip(formula.get("rhs", []), rhs_indices):
                                rhs_parts_r.append(f"{_val(r, idx):.2f}")
                            fail_samples.append(
                                f"{_row_name(r)}: {lhs_val_r:.2f}≠{' + '.join(rhs_parts_r)}={rhs_sum_r:.2f}")
                if fail_count == 0:
                    _add_check(formula_name + " (逐行)", "per_row", True,
                                f"{pass_count} 行全部通过")
                else:
                    sample_str = " | ".join(fail_samples)
                    _add_check(formula_name + " (逐行)", "per_row", False,
                                f"通过 {pass_count} 行，失败 {fail_count} 行 | {sample_str}")

    # ──────────────────────────────────────────
    # Column sum checks (列加总校验)
    # ──────────────────────────────────────────
    col_sum_failures = 0
    for target_kw_list in col_sum_targets:
        idx = _col(target_kw_list)
        if idx < 0:
            continue
        col_total = round(sum(_val(r, idx) for r in range(data_start, data_end + 1)), 2)
        ref_val = _val(summary_row, idx)
        diff = round(col_total - ref_val, 2)
        label = target_kw_list[0]
        if abs(diff) > tol:
            col_sum_failures += 1
            _add_check(f"{label} 列加总", "column_sum", False,
                        f"数据行合计 {col_total:.2f} ≠ 合计行 {ref_val:.2f} (差 {diff:+.2f})")

    if col_sum_failures == 0 and col_sum_targets:
        _add_check("列加总校验", "column_sum", True)

    # ──────────────────────────────────────────
    # Critical column checks (关键列存在性)
    # ──────────────────────────────────────────
    if critical_columns:
        missing = []
        for label, kw_list in critical_columns.items():
            if _col(kw_list if isinstance(kw_list, list) else [kw_list]) < 0:
                missing.append(label)
        if missing:
            _add_check("关键列存在性", "table_format", False,
                        f"缺失: {'、'.join(missing)}")
        else:
            _add_check("关键列存在性", "table_format", True)

    passed_count = sum(1 for c in checks if c["passed"])
    failed_count = len(checks) - passed_count

    result = {
        "ok": failed_count == 0,
        "passed_count": passed_count,
        "failed_count": failed_count,
        "checks": checks,
    }

    # ── 追加验证 sheet ──
    _append_validation_sheet(wb, checks)
    wb.save(filepath)
    wb.close()
    return result


def _append_validation_sheet(wb, checks):
    """
    在现有 workbook 中追加「验证结果」sheet，写入校验明细。
    checks: [{"name", "kind", "passed", "detail"}, ...]
    """
    from openpyxl.styles import PatternFill, Font, Alignment

    ws = wb.create_sheet(title="验证结果")

    # Styles
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=12)

    # Title
    ws.cell(row=1, column=1, value="工资表数据验证结果")
    ws.cell(row=1, column=1).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    # Headers
    headers = ["#", "校验项", "结果", "说明"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    for i, chk in enumerate(checks, 1):
        row = i + 2
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=chk["name"])
        status = "✅ 通过" if chk["passed"] else "❌ 未通过"
        ws.cell(row=row, column=3, value=status)
        ws.cell(row=row, column=4, value=chk.get("detail", ""))

        fill = green_fill if chk["passed"] else red_fill
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = fill

    # Summary row
    summary_row = len(checks) + 3
    ws.cell(row=summary_row, column=2, value="汇总")
    ws.cell(row=summary_row, column=2).font = bold_font
    passed = sum(1 for c in checks if c["passed"])
    failed = len(checks) - passed
    ws.cell(row=summary_row, column=3, value=f"✅ {passed} 项通过" if failed == 0 else f"⚠️ {passed} 项通过, {failed} 项未通过")
    if failed == 0:
        ws.cell(row=summary_row, column=3).font = Font(bold=True, color="006100")
    else:
        ws.cell(row=summary_row, column=3).font = Font(bold=True, color="9C0006")

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 60


def _log_validation_results(gui, validation_results):
    """向 GUI 日志输出验证结果汇总。"""
    for label in ("工资表",):
        vr = validation_results.get(label)
        if vr:
            if vr["ok"]:
                gui.log(f"  ✅ {label}：全部 {vr['passed_count']} 项校验通过")
                for chk in vr["checks"]:
                    if not chk["passed"]:
                        continue
                    gui.log(f"    ✓ {chk['name']}")
            else:
                gui.log(f"  ⚠️ {label}：{vr['passed_count']}/{vr['passed_count'] + vr['failed_count']} 项通过")
                for chk in vr["checks"]:
                    if chk["passed"]:
                        gui.log(f"    ✓ {chk['name']}")
                    else:
                        detail = f" — {chk['detail']}" if chk.get("detail") else ""
                        gui.log(f"    ✗ {chk['name']}{detail}")
        elif label in validation_results:
            gui.log(f"  - {label}：未生成文件")


def merge_payrolls_by_tax(payroll_dir, output_dir, bank_dir=None, selected_title=None):
    """
    合并所有工资表为一张表，不再按个税分表。
    智能命名：不同结算单元→{名}{count}家{月}，相同→{名}{月}。
    同时生成对应的银行报盘文件。
    selected_title: 用户选择的输出标题（不含年月替换），None 则自动取最众数
    返回: (payroll_path, bank_path, validation_results, op_log_path)
    """
    from collections import defaultdict
    output_dir = os.path.normpath(output_dir)
    if bank_dir:
        bank_dir = os.path.normpath(bank_dir)
    payroll_dir = os.path.normpath(payroll_dir)
    os.makedirs(output_dir, exist_ok=True)

    # 扫描工资表
    payroll_files = [
        f for f in os.listdir(payroll_dir)
        if "汇总表" not in f and "验证" not in f
        and f.endswith((".xlsx", ".xls"))
    ]

    # 去重排序：同一单位优先保留 signed_ 版本，其次 xlsx，最后 xls
    unit_file_map = {}
    for fname in payroll_files:
        unit_name = fname
        for sep in ["2026年06月工资表", "202606工资表", "202606工资"]:
            if sep in fname:
                unit_name = fname.split(sep, 1)[0].strip()
                break
        unit_name = unit_name.replace("signed_", "").strip()
        is_signed = fname.startswith("signed_")
        is_xlsx = fname.endswith(".xlsx")
        priority = 0 if is_signed and is_xlsx else 1 if is_signed else 2 if is_xlsx else 3
        if unit_name not in unit_file_map or priority < unit_file_map[unit_name][0]:
            unit_file_map[unit_name] = (priority, fname)

    sorted_files = sorted(unit_file_map.values(), key=lambda x: x[0])

    # 读取所有工资表
    all_data = []  # (unit_name, tax_amt, row, headers, fname, source_excel_row)
    all_file_col_names = []
    file_name_to_idx_map = {}
    sample_header = None
    sample_merged_cells = None
    sample_images = None
    fill_dates = []  # 各文件填报时间
    maker_names = []  # 各文件制表人（从表尾签字行提取）
    file_years_months = []  # 各文件标题中的年月（用于输出标题）
    file_titles = []  # 各文件标题行原文

    for priority, fname in sorted_files:
        if not fname.startswith("signed_"):
            continue  # 未签字工资表不参与合并
        path = os.path.join(payroll_dir, fname)
        headers, data_rows, footers, tax_col = _read_payroll_data(path)

        name_to_idx, col_names = _build_column_name_map(headers)
        all_file_col_names.append(col_names)
        file_name_to_idx_map[fname] = name_to_idx

        if sample_header is None:
            sample_header = headers

        # 扫描签名图片（第一个有图片的 signed 文件）
        if not sample_images and fname.startswith("signed_") and fname.endswith(".xlsx"):
            try:
                _, ws, _, _, _, merged, images = _read_payroll_workbook(path)
                if images:
                    sample_images = images
            except Exception:
                pass

        unit_name = ""
        file_ym = ""
        if len(headers) > 1:
            raw = headers[1]
            # B2（column 1）为实际单位名称；若为空/仅"单位"，回退 A2
            val = str(raw[1]).strip() if len(raw) > 1 and raw[1] else ""
            if not val or val in ("单位",):
                val = str(raw[0]).strip() if len(raw) > 0 and raw[0] else ""
            for pfx in ("单位名称：", "名称："):
                if val.startswith(pfx):
                    val = val[len(pfx):]
                    break
            unit_name = val.strip()
            # 收集填报时间（仅用于操作记录）
            for cell_val in raw:
                if cell_val and "填报时间" in str(cell_val):
                    import re
                    m = re.search(r"\d{4}-\d{2}-\d{2}", str(cell_val))
                    if m:
                        fill_dates.append(m.group(0))
            # 从标题行提取年月，并收集标题原文用作输出模板
            if headers and len(headers) > 0:
                import re
                for cell_val in headers[0]:
                    _s = str(cell_val or "")
                    if _s and _s not in ("None",):
                        file_titles.append(_s)
                    _m = re.search(r"(\d{4})年(\d{1,2})月", _s)
                    if _m:
                        file_ym = f"{_m.group(1)}年{int(_m.group(2)):02d}月"
                        if file_ym:
                            file_years_months.append(file_ym)
                        break
            # 收集制表人
            import re as _re
            for _row in (list(headers) + list(footers)):
                for _cv in _row:
                    _s = str(_cv or "").strip()
                    if _s.startswith("制表人") and "：" in _s:
                        _name = _s.split("：", 1)[1].strip()
                        if _name:
                            maker_names.append(_name)

        if not unit_name:
            # 兜底：从文件名提取
            unit_name = fname
            for sep in ["2026年06月工资表", "202606工资表", "202606工资"]:
                if sep in fname:
                    unit_name = fname.split(sep, 1)[0].strip()
                    break
            unit_name = unit_name.replace("signed_", "").strip()

        for row_idx, row in enumerate(data_rows):
            tax_val = row[tax_col] if tax_col is not None else 0
            try:
                tax_amt = float(tax_val) if tax_val else 0
            except (ValueError, TypeError):
                tax_amt = 0
            source_excel_row = row_idx + 6  # 5 行表头 + 1-based
            all_data.append((unit_name, tax_amt, row, headers, fname, source_excel_row, file_ym))

    if not all_data:
        return None, None, {}, None

    # 取所有源文件中出现次数最多的填报时间
    from collections import Counter
    fill_date_counts = Counter(fill_dates)
    most_common_fill_date = fill_date_counts.most_common(1)[0][0] if fill_date_counts else ""

    canonical_cols = _get_canonical_columns(all_file_col_names)
    canonical_cols = [c for c in canonical_cols if c not in ("部门", "岗位", "职工号")]

    normalized_data = []  # (unit_name, tax_amt, norm_row, source_file, source_excel_row, file_ym)
    for unit_name, tax_amt, row, headers, fname, source_excel_row, file_ym in all_data:
        norm_row = _normalize_row_by_names(row, headers, canonical_cols)
        normalized_data.append((unit_name, tax_amt, norm_row, fname, source_excel_row, file_ym))

    name_col_in_canonical = None
    for c, name in enumerate(canonical_cols):
        if name == "姓名":
            name_col_in_canonical = c
            break
    if name_col_in_canonical is not None:
        insert_pos = name_col_in_canonical + 1
        canonical_cols.insert(insert_pos, "结算单元")
        for i, rec in enumerate(normalized_data):
            uname, tax_amt, nrow, fname, src_row, file_ym = rec
            nrow.insert(insert_pos, uname)
            normalized_data[i] = (uname, tax_amt, nrow, fname, src_row, file_ym)

    # "月份" 列：紧随结算单元之后
    settle_unit_pos = canonical_cols.index("结算单元") if "结算单元" in canonical_cols else -1
    if settle_unit_pos >= 0:
        month_pos = settle_unit_pos + 1
        canonical_cols.insert(month_pos, "月份")
        for i, rec in enumerate(normalized_data):
            uname, tax_amt, nrow, fname, src_row, file_ym = rec
            nrow.insert(month_pos, file_ym if file_ym else "")
            normalized_data[i] = (uname, tax_amt, nrow, fname, src_row, file_ym)

    # 删除全零列（从第4列 基本工资 开始检查）
    cols_to_remove = set()
    for c in range(4, len(canonical_cols)):
        all_zero = True
        for rec in normalized_data:
            row = rec[2]
            val = row[c] if c < len(row) else None
            try:
                if float(val or 0) != 0:
                    all_zero = False
                    break
            except (ValueError, TypeError):
                if val and str(val).strip():
                    all_zero = False
                    break
        if all_zero:
            cols_to_remove.add(c)

    if cols_to_remove:
        for c in sorted(cols_to_remove, reverse=True):
            del canonical_cols[c]
            for i in range(len(normalized_data)):
                del normalized_data[i][2][c]

    # 确保 大病险 在 转账合计 左侧
    if "转账合计" in canonical_cols and "大病险" in canonical_cols:
        trans_idx = canonical_cols.index("转账合计")
        illness_idx = canonical_cols.index("大病险")
        if illness_idx > trans_idx:
            canonical_cols.pop(illness_idx)
            canonical_cols.insert(trans_idx, "大病险")
            for i in range(len(normalized_data)):
                rec = list(normalized_data[i])
                val = rec[2].pop(illness_idx)
                rec[2].insert(trans_idx, val)
                normalized_data[i] = tuple(rec)
            # 调整 max_output_cols（后面会重新从 canonical_cols 长度取）
            # 转账合计公式用的大病险值不变，只是列移动了

    # 统计结算单元数量用于智能命名
    settle_unit_col = next((i for i, n in enumerate(canonical_cols) if n == "结算单元"), None)
    settle_unit_candidates = set()
    for rec in normalized_data:
        su = rec[2][settle_unit_col] if settle_unit_col is not None and settle_unit_col < len(rec[2]) else ""
        if su:
            settle_unit_candidates.add(su)
    unique_unit_count = len(settle_unit_candidates)

    # 从各文件标题取所有月份，生成涵盖所有月份的标题
    import re as _re
    _all_ym = set()
    for _ym in file_years_months:
        _m3 = _re.match(r"(\d{4})年(\d{2})月", _ym)
        if _m3:
            _all_ym.add((int(_m3.group(1)), int(_m3.group(2))))
    _sorted_ym = sorted(_all_ym)

    if _sorted_ym:
        # 按年分组
        from itertools import groupby
        _title_parts = []
        _all_month_nums = []
        for _yr, _grp in groupby(_sorted_ym, key=lambda x: x[0]):
            _mons = sorted(m for _, m in _grp)
            _all_month_nums.extend(_mons)
            if len(_mons) == 1:
                _title_parts.append(f"{_yr}年{_mons[0]:02d}月")
            elif all(_mons[i] + 1 == _mons[i+1] for i in range(len(_mons)-1)):
                _title_parts.append(f"{_yr}年{_mons[0]:02d}-{_mons[-1]:02d}月")
            else:
                _title_parts.append(f"{_yr}年{'、'.join(str(m) for m in _mons)}月")
        year_month = "".join(_title_parts)

        # 文件名用短格式
        if len(_sorted_ym) == 1:
            month_short = f"{_sorted_ym[0][1]}月"
        else:
            _flat = sorted(set(_all_month_nums))
            if all(_flat[i] + 1 == _flat[i+1] for i in range(len(_flat)-1)):
                month_short = f"{_flat[0]}-{_flat[-1]}月"
            else:
                month_short = '_'.join(str(m) for m in _flat) + "月"
    else:
        year_month = "2026年06月"
        month_short = "6月"

    # 从结算单元名称中提取公共前缀作为单位基名
    def _common_prefix(strings):
        if not strings:
            return ""
        first = min(strings)
        last = max(strings)
        for i, (a, b) in enumerate(zip(first, last)):
            if a != b:
                return first[:i]
        return first

    settle_unit_names = sorted(settle_unit_candidates)
    org_prefix = _common_prefix(settle_unit_names) if settle_unit_names else "吉林大学"

    # 智能命名：不同结算单元→{基名}{count}家{月}，相同→{基名}{月}
    if unique_unit_count > 1:
        base_name = f"{org_prefix}{unique_unit_count}家{month_short}"
    else:
        base_name = f"{org_prefix}{month_short}"

    payroll_fname = f"{base_name}工资表.xlsx"
    bank_fname = f"{base_name}报盘.xlsx"

    all_group = normalized_data

    max_output_cols = len(canonical_cols)

    header_rows = []

    # 输出标题：用户选定的 → 替换年月；未选定 → 自动构造
    import re
    if selected_title:
        output_title = re.sub(r'\d{4}年\d{1,2}月', year_month, selected_title)
    elif file_titles:
        # 取最众数标题中的后缀（年月之后的部分），与 org_prefix + year_month 拼接
        from collections import Counter
        _common_title = Counter(file_titles).most_common(1)[0][0]
        _re_suffix = re.search(r'\d{4}年\d{1,2}月(.+)$', _common_title)
        if _re_suffix:
            output_title = f"{org_prefix}{year_month}{_re_suffix.group(1)}"
        else:
            output_title = f"{org_prefix}{year_month}人才派遣人员工资发放表"
    else:
        output_title = f"{org_prefix}{year_month}人才派遣人员工资发放表"
    title_row = [""] * max_output_cols
    title_row[0] = output_title
    header_rows.append(title_row)

    unit_row = [""] * max_output_cols
    unit_row[0] = f"单位名称：{org_prefix}"
    header_rows.append(unit_row)

    main_row = []
    sub1_row = []
    sub2_row = []
    for cname in canonical_cols:
        if cname == "结算单元":
            main_row.append("结算单元")
            sub1_row.append("")
            sub2_row.append("")
            continue
        if "/" in cname:
            parts = cname.split("/", 1)
            main_val = parts[0]
            main_row.append(main_val)
            sub = parts[1]
            if "/" in sub:
                sp = sub.split("/", 1)
                sub1_row.append(sp[0])
                sub2_row.append(sp[1])
            else:
                sub1_row.append(sub)
                sub2_row.append("")
        else:
            main_row.append(cname)
            sub1_row.append("")
            sub2_row.append("")

    header_rows.append(main_row)
    if any(v for v in sub1_row):
        header_rows.append(sub1_row)
    if any(v for v in sub2_row):
        header_rows.append(sub2_row)

    # 生成合并单元格列表（非连续的扣款明细列拆分为多个合并范围）
    # 找到所有连续的扣款明细列区间
    detail_ranges = []
    range_start = None
    for c, cname in enumerate(canonical_cols, 1):
        if cname.startswith("扣款明细"):
            if range_start is None:
                range_start = c
        else:
            if range_start is not None:
                if c - 1 > range_start:
                    detail_ranges.append((range_start, c - 1))
                range_start = None
    if range_start is not None and len(canonical_cols) >= range_start:
        detail_ranges.append((range_start, len(canonical_cols)))

    title_end_col = openpyxl.utils.get_column_letter(max_output_cols)

    canonical_merged = []
    canonical_merged.append(f"A1:{title_end_col}1")
    for ds, de in detail_ranges:
        ds_col = openpyxl.utils.get_column_letter(ds)
        de_col = openpyxl.utils.get_column_letter(de)
        canonical_merged.append(f"{ds_col}3:{de_col}3")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    def write_payroll(group, fname):
        if not group:
            return None, []
        fpath = os.path.join(output_dir, fname)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "工资发放表"

        # ── 页面设置 ──
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        from openpyxl.worksheet.properties import PageSetupProperties
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.paperSize = 9  # A4
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.print_title_rows = '1:5'  # 每页重复表头

        # ── 通用样式 ──
        thin_side = Side(style='thin')
        thin_border = Border(left=thin_side, right=thin_side,
                             top=thin_side, bottom=thin_side)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        center_align_nowrap = Alignment(horizontal='center', vertical='center')
        header_font = Font(bold=True, size=10)
        title_font = Font(bold=True, size=14)

        num_header_rows = len(header_rows)  # 3 或 4

        # ── 写入表头 ──
        for r_idx, row in enumerate(header_rows):
            for c_idx, val in enumerate(row):
                cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=val)
                if r_idx >= 2:  # 第 1、2 行（标题/单位）无边框
                    cell.border = thin_border
                cell.alignment = center_align
                cell.font = header_font

        # 第一行标题行：大字号、跨列合并已在 canonical_merged
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = title_font

        # ── 写入数据行 ──
        data_start = num_header_rows + 1
        settle_unit_col = next((i for i, n in enumerate(canonical_cols) if n == "结算单元"), None)
        # 记录每个源数据行对应的输出行号
        output_row_map = []  # [(source_file, source_excel_row, output_excel_row), ...]
        for i, rec in enumerate(group, 1):
            uname, tax_amt, row, fname, src_row = rec[0], rec[1], rec[2], rec[3], rec[4] if len(rec) > 3 else ("", 0)
            output_row_map.append((fname, src_row, data_start + i - 1))
            row[0] = i
            for c_idx, val in enumerate(row):
                cell = ws.cell(row=data_start + i - 1, column=c_idx + 1, value=val)
                cell.border = thin_border
                cell.alignment = (Alignment(horizontal='center', vertical='center', wrap_text=True)
                                  if c_idx == settle_unit_col else center_align_nowrap)
                if c_idx >= 6:
                    try:
                        float(val)
                        cell.number_format = "0.00"
                    except (ValueError, TypeError):
                        pass

        # ── 合计行 ──
        total_row_idx = num_header_rows + len(group) + 1
        cell = ws.cell(row=total_row_idx, column=1, value="合计")
        cell.border = thin_border
        cell.alignment = center_align
        cell.font = Font(bold=True, size=10)
        for c in range(2, max_output_cols + 1):
            total = 0.0
            all_num = True
            for rec in group:
                row = rec[2]
                v = row[c - 1] if c - 1 < len(row) else 0
                try:
                    total += float(v or 0)
                except (ValueError, TypeError):
                    all_num = False
                    break
            if all_num:
                cell = ws.cell(row=total_row_idx, column=c, value=total)
                cell.number_format = "0.00"
            else:
                cell = ws.cell(row=total_row_idx, column=c)
            cell.border = thin_border
            cell.alignment = center_align
            cell.font = Font(bold=True, size=10)

        # ── 签名行 ──
        sign_row_idx = total_row_idx + 3
        # 列出所有不重复制表人
        _unique_makers = []
        for _n in maker_names:
            if _n not in _unique_makers:
                _unique_makers.append(_n)
        _maker_text = "制表人：" + "、".join(_unique_makers) if _unique_makers else "制表人："
        sign_labels = {
            1: "总经理签字：",
            7: "部长签字：",
            13: "财务审核：",
            18: "业务审核：",
            24: _maker_text,
        }
        for col, label in sign_labels.items():
            cell = ws.cell(row=sign_row_idx, column=col, value=label)
            if col != 24:  # 制表人无需换行，允许超出
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # ── 合并单元格（按规则） ──
        from collections import defaultdict

        # 标题行 A1:AH1
        ws.merge_cells(f"A1:{title_end_col}1")

        # Row 2（单位信息）：左侧单位名称 + 右侧填报时间
        split_col = min(22, max_output_cols)
        left_end = openpyxl.utils.get_column_letter(split_col)
        try:
            ws.merge_cells(f"A2:{left_end}2")
        except Exception:
            pass
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='left', vertical='center')
        # 右侧合并最后 6 列显示填报时间（避免相邻空字符串阻挡文字溢出）
        right_start = max(split_col + 1, max_output_cols - 5)
        date_start_letter = openpyxl.utils.get_column_letter(right_start)
        fill_date_str = f"填报时间：{most_common_fill_date}" if most_common_fill_date else ""
        if fill_date_str:
            # 先清空，再合并
            for c in range(right_start, max_output_cols + 1):
                ws.cell(row=2, column=c).value = None
            try:
                ws.merge_cells(f"{date_start_letter}2:{title_end_col}2")
            except Exception:
                pass
            ws.cell(row=2, column=right_start, value=fill_date_str
                   ).alignment = Alignment(horizontal='right', vertical='center')

        # Rows 3-5：按 canonical_cols 规则合并

        # 收集单层列（不含 "/"）→ {col}3:{col}5 垂直合并
        single_level_cols = []
        # 收集双层列（仅 2 段，如 扣款明细/单位代理费）→ {col}4:{col}5 垂直合并
        two_level_cols = []
        # Row 3 分组：相同一级名的列 → 水平合并
        main_groups = defaultdict(list)  # main_name → [col_indices]
        # Row 4 分组：相同二级名的列 → 水平合并（仅扣款明细区域）
        sub1_groups = defaultdict(list)  # sub1_name → [col_indices]

        for c_idx, cname in enumerate(canonical_cols, 1):
            if "/" not in cname:
                single_level_cols.append(c_idx)
                main_groups[cname].append(c_idx)
            else:
                parts = cname.split("/", 1)
                main_groups[parts[0]].append(c_idx)
                sub = parts[1]
                if "/" in sub:
                    sub1_groups[sub.split("/")[0]].append(c_idx)
                else:
                    sub1_groups[sub].append(c_idx)
                    two_level_cols.append(c_idx)

        # Row 3：水平合并同组多列
        for group_cols in main_groups.values():
            if len(group_cols) > 1:
                cs = openpyxl.utils.get_column_letter(group_cols[0])
                ce = openpyxl.utils.get_column_letter(group_cols[-1])
                try:
                    ws.merge_cells(f"{cs}3:{ce}3")
                except Exception:
                    pass

        # Row 4：水平合并同组多列（sub1）
        for group_cols in sub1_groups.values():
            if len(group_cols) > 1:
                cs = openpyxl.utils.get_column_letter(group_cols[0])
                ce = openpyxl.utils.get_column_letter(group_cols[-1])
                try:
                    ws.merge_cells(f"{cs}4:{ce}4")
                except Exception:
                    pass

        # 单层列垂直合并 {col}3:{col}5
        for c_idx in single_level_cols:
            cl = openpyxl.utils.get_column_letter(c_idx)
            try:
                ws.merge_cells(f"{cl}3:{cl}5")
            except Exception:
                pass

        # 双层列垂直合并 {col}4:{col}5
        for c_idx in two_level_cols:
            cl = openpyxl.utils.get_column_letter(c_idx)
            try:
                ws.merge_cells(f"{cl}4:{cl}5")
            except Exception:
                pass

        # ── 行高 ──
        ws.row_dimensions[1].height = 36  # 标题行
        for rr in range(2, num_header_rows + 1):
            ws.row_dimensions[rr].height = 28  # 表头行
        ws.row_dimensions[total_row_idx].height = 24  # 合计行

        # ── 列宽（根据内容类型） ──
        for c, cname in enumerate(canonical_cols, 1):
            if cname in ("序号",):
                ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 6
            elif cname in ("姓名",):
                ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 10
            elif cname in ("身份证",):
                ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 20
            elif cname in ("结算单元",):
                # 动态计算列宽：基于数据实际宽度，取 90% 分位
                widths = []
                for rec in group:
                    row = rec[2]
                    val = str(row[c - 1]) if c - 1 < len(row) else ""
                    if not val:
                        continue
                    # CJK 字符约 2 个宽度单位，ASCII 约 1
                    w = sum(2 if ord(ch) > 127 else 1 for ch in val)
                    widths.append(w)
                if widths:
                    widths.sort()
                    p90 = widths[int(len(widths) * 0.9)]
                    col_w = max(min(p90 + 2, 30), 6)
                else:
                    col_w = 12
                ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = col_w
            elif cname.startswith("扣款明细"):
                ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 9
            else:
                # 数值列：取数据最大宽度 + 合计行宽度
                max_w = 11
                max_val_w = 0
                total_val = 0.0
                has_total = False
                for rec in group:
                    row = rec[2]
                    v = row[c - 1] if c - 1 < len(row) else 0
                    try:
                        num = float(v or 0)
                        val_str = f"{num:.2f}"
                        val_w = len(val_str)
                        max_val_w = max(max_val_w, val_w)
                        total_val += num
                        has_total = True
                    except (ValueError, TypeError):
                        val_str = str(v) if v else ""
                        max_val_w = max(max_val_w, sum(2 if ord(ch) > 127 else 1 for ch in val_str))
                if has_total:
                    total_str = f"{total_val:.2f}"
                    max_w = max(max_val_w, len(total_str)) + 3
                else:
                    max_w = max(max_val_w, 11) + 3
                max_w = min(max(max_w, 8), 20)
                ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = max_w

        # ── 签名图片（按左侧签名字段右移一列放置）──
        if sample_images:
            from openpyxl.drawing.image import Image as XLImage
            import io
            # 签名字段列 → 图片锚点列（右移一列，文字自动换行不溢出故不重叠）
            sig_cols = {1: 2, 7: 8, 13: 14, 18: 19}
            # 按原始列升序排列，保证左→右对应关系
            sorted_imgs = sorted(sample_images, key=lambda x: x[2])
            TARGET_IMG_H = 64  # px, ≈3个行高高度
            for (img_data, _orig_row, _orig_col, width, height), out_col in zip(sorted_imgs, sig_cols.values()):
                try:
                    new_img = XLImage(io.BytesIO(img_data))
                    scale = TARGET_IMG_H / height if height else 1.0
                    new_img.width = int(width * scale)
                    new_img.height = TARGET_IMG_H
                    cell_ref = openpyxl.utils.get_column_letter(out_col) + str(sign_row_idx)
                    new_img.anchor = cell_ref
                    ws.add_image(new_img)
                except Exception:
                    pass
            ws.row_dimensions[sign_row_idx].height = 50  # pt, 容纳签名图

        wb.save(fpath)
        return fpath, output_row_map

    payroll_result = write_payroll(all_group, payroll_fname)
    payroll_path = payroll_result[0]
    payroll_row_map = payroll_result[1] if payroll_result else []

    # ── 验证 ──
    validation_results = {}
    if payroll_path:
        try:
            validation_results["工资表"] = validate_payroll_xlsx(payroll_path, canonical_cols)
        except Exception as e:
            validation_results["工资表"] = {"ok": False, "passed_count": 0, "failed_count": 1,
                                           "checks": [{"name": "验证执行异常", "passed": False, "detail": str(e)}]}

    # 生成对应的银行报盘文件
    bank_path = None
    bank_prov = []
    if bank_dir:
        bank_files = [f for f in os.listdir(bank_dir) if f.lower().endswith(".xls")]
        bank_map = defaultdict(list)
        for fname in bank_files:
            try:
                _, _, unit_name = split_filename(fname)
                bank_map[unit_name].append(fname)
            except ValueError:
                continue

        _script_dir = os.path.dirname(os.path.abspath(__file__))
        bank_tmpl_path = os.path.join(_script_dir, "template", "代发业务导入模板.xlsx")

        def write_bank(group_data, fname):
            if not group_data:
                return None, []
            all_bank_rows = []  # [(row_data, source_file, source_row, bank_type), ...]
            seen_units = set()
            for rec in group_data:
                unit_name = rec[0]
                if unit_name in bank_map and unit_name not in seen_units:
                    seen_units.add(unit_name)
                    for bfname in bank_map[unit_name]:
                        bfpath = os.path.join(bank_dir, bfname)
                        bt = detect_bank_type(bfname)
                        if bt == "icbc":
                            rows = _read_icbc_rows(bfpath, [])
                        elif bt == "ccb":
                            rows = _read_ccb_rows(bfpath, [])
                        elif bt == "jlb":
                            rows = _read_jlb_rows(bfpath, [])
                        else:
                            continue
                        for src_idx, row in enumerate(rows):
                            src_excel_row = src_idx + 2
                            all_bank_rows.append((row, bfname, src_excel_row, bt))

            if not all_bank_rows:
                return None, []

            for i, (row, _, _, _) in enumerate(all_bank_rows, 1):
                row[0] = i

            fpath = os.path.join(output_dir, fname)
            tmpl = openpyxl.load_workbook(bank_tmpl_path)
            ws = tmpl["代发工资模板"]
            ws.title = "银行报盘"
            for r in range(5, 8):
                for c in range(1, 7):
                    ws.cell(row=r, column=c).value = None
            prov = []
            for out_idx, (row, src_file, src_row, bt) in enumerate(all_bank_rows, start=5):
                ws.cell(row=out_idx, column=1, value=out_idx - 4)
                ws.cell(row=out_idx, column=2, value=str(row[1]).strip())
                ws.cell(row=out_idx, column=3, value=str(row[2]).strip())
                cell_e = ws.cell(row=out_idx, column=5, value=row[3])
                cell_e.number_format = "0.00"
                prov.append((src_file, src_row,
                             str(row[1]) if len(row) > 1 else "",
                             str(row[2]) if len(row) > 2 else "",
                             str(row[3]) if len(row) > 3 else "",
                             bt,
                             fname, out_idx - 4))
            last_data_row = 4 + len(all_bank_rows)
            # 用计算值替代公式，避免打开时触发保存提示
            total_amount = sum(float(row[3] or 0) for (row, _, _, _) in all_bank_rows)
            ws["B2"] = total_amount
            ws["B2"].number_format = "0.00"
            ws["B3"] = len(all_bank_rows)
            tmpl.save(fpath)
            return fpath, prov

        bank_result = write_bank(all_group, bank_fname)
        bank_path = bank_result[0] if bank_result else None
        bank_prov = bank_result[1] if bank_result else []

        # 重命名文件：追加人数/笔数和总金额
        # 从 canonical_cols 找实发合计的列索引
        pay_col = next((i for i, n in enumerate(canonical_cols) if n == "实发合计"), None)
        payroll_total = 0.0
        if pay_col is not None:
            for rec in all_group:
                nrow = rec[2]
                try:
                    payroll_total += float(nrow[pay_col] or 0)
                except (ValueError, TypeError, IndexError):
                    pass

        bank_total = 0.0
        for bp in bank_prov:
            try:
                bank_total += float(bp[4] or 0)
            except (ValueError, TypeError):
                pass

        new_payroll = f"{base_name}工资表_{len(payroll_row_map)}人_{payroll_total:.2f}元.xlsx"
        new_bank = f"{base_name}报盘_{len(bank_prov)}笔_{bank_total:.2f}元.xlsx"

        if payroll_path:
            new_payroll_path = os.path.join(output_dir, new_payroll)
            try:
                os.replace(payroll_path, new_payroll_path)
                payroll_path = new_payroll_path
                payroll_fname = new_payroll
            except OSError:
                pass
        if bank_path:
            new_bank_path = os.path.join(output_dir, new_bank)
            try:
                os.replace(bank_path, new_bank_path)
                bank_path = new_bank_path
                bank_fname = new_bank
            except OSError:
                pass

    # ── 构建操作记录 ──
    col_name_idx = {n: i for i, n in enumerate(canonical_cols)}
    payroll_provenance = []  # [(src_file, src_row, unit, name, id_num, total_pay, out_file, out_row, settle_unit), ...]
    for out_row, rec in enumerate(payroll_row_map, 1):
        fname_src, src_row, _ = rec
        # 找到对应的原始数据记录
        data_rec = all_group[out_row - 1]
        uname, _, nrow, src_file, src_row2, _ = data_rec
        name = str(nrow[col_name_idx["姓名"]]) if "姓名" in col_name_idx and col_name_idx["姓名"] < len(nrow) else ""
        id_no = str(nrow[col_name_idx["身份证"]]) if "身份证" in col_name_idx and col_name_idx["身份证"] < len(nrow) else ""
        total_pay = nrow[col_name_idx["实发合计"]] if "实发合计" in col_name_idx and col_name_idx["实发合计"] < len(nrow) else ""
        settle_unit = str(nrow[col_name_idx["结算单元"]]) if "结算单元" in col_name_idx and col_name_idx["结算单元"] < len(nrow) else ""
        payroll_provenance.append((src_file, src_row, uname, name, id_no, total_pay, payroll_fname, out_row, settle_unit))

    # 姓名 → 银行记录索引（一对多，重名可能有多个）
    bank_by_name = defaultdict(list)
    for b_idx, bp in enumerate(bank_prov):
        bank_name = bp[3].strip()  # 户名
        if bank_name:
            bank_by_name[bank_name].append(b_idx)

    # 生成操作记录 Excel
    op_log_path = os.path.join(output_dir, f"操作记录_{ts}.xlsx")
    log_wb = openpyxl.Workbook()

    # ---- Sheet 1: 工资表与银行报盘（合并） ----
    ws1 = log_wb.active
    ws1.title = "工资表与银行报盘"
    ws1.append(["原始工资表文件", "原始行号", "原始单位名称", "姓名", "身份证号", "实发合计",
                "目标单位（结算单元）", "银行账号", "银行户名", "银行金额", "银行类型",
                "原始报盘文件", "匹配状态", "金额一致", "备注"])
    # 已匹配 / 未匹配的工资表记录
    for pp in payroll_provenance:
        src_file, src_row, unit, name, id_no, total_pay, out_file, out_row, settle_unit = pp
        matched_banks = bank_by_name.get(name.strip(), [])
        if matched_banks:
            bi = matched_banks[0]
            bp = bank_prov[bi]
            acct = bp[2]       # 账号
            acct_name = bp[3]  # 户名
            amount = bp[4]     # 金额
            bt = bp[5]         # 银行类型
            bsrc = bp[0]       # 原始报盘文件
            # 金额一致性校验
            try:
                pay_val = float(total_pay or 0)
                bank_val = float(amount or 0)
                if abs(pay_val - bank_val) < 0.01:
                    amt_ok = "一致"
                else:
                    amt_ok = f"差{pay_val - bank_val:.2f}"
            except (ValueError, TypeError):
                amt_ok = "无法比较"
            if len(matched_banks) > 1:
                extra = "; ".join(f"{bank_prov[b][5]}/{bank_prov[b][4]}元" for b in matched_banks[1:])
                status = "重名"
                note = f"另有匹配: {extra}"
            else:
                status = "已匹配"
                note = ""
        else:
            acct = acct_name = amount = bt = bsrc = ""
            status = "未匹配"
            amt_ok = ""
            note = "无对应银行记录"
        ws1.append([src_file, src_row, unit, name, id_no, total_pay,
                    settle_unit, acct, acct_name, amount, bt, bsrc, status, amt_ok, note])

    # 银行有但工资表没有的报盘记录
    payroll_names = set(pp[3].strip() for pp in payroll_provenance)
    for b_idx, bp in enumerate(bank_prov):
        bank_name = bp[3].strip()
        if bank_name and bank_name not in payroll_names:
            ws1.append(["", "", "", bank_name, "", "",
                        "", bp[2], bp[3], bp[4], bp[5], bp[0], "银行无对应工资表", ""])

    # ---- Sheet 2: 汇总 ----
    ws2 = log_wb.create_sheet("汇总")
    ws2.append(["指标", "值"])
    ws2.append(["填报时间", most_common_fill_date])
    ws2.append(["操作时间", ts])
    ws2.append(["结算单元数", unique_unit_count])
    ws2.append(["工资表总人数", len(payroll_provenance)])
    ws2.append(["银行报盘总笔数", len(bank_prov)])
    ws2.append(["银行已匹配人数", sum(1 for pp in payroll_provenance if bank_by_name.get(pp[3].strip(), []))])
    ws2.append(["银行未匹配人数", sum(1 for pp in payroll_provenance if not bank_by_name.get(pp[3].strip(), []))])
    ws2.append(["重名人数", sum(1 for bps in bank_by_name.values() if len(bps) > 1)])
    # 总金额校验
    try:
        total_pay = sum(float(pp[5] or 0) for pp in payroll_provenance)
        total_bank = sum(float(bp[4] or 0) for bp in bank_prov)
        ws2.append(["工资表总金额", f"{total_pay:.2f}"])
        ws2.append(["银行报盘总金额", f"{total_bank:.2f}"])
        ws2.append(["金额差异", f"{total_pay - total_bank:.2f}"])
    except (ValueError, TypeError):
        pass
    ws2.append(["输出文件（工资表）", os.path.basename(payroll_path) if payroll_path else ""])
    ws2.append(["输出文件（报盘）", os.path.basename(bank_path) if bank_path else ""])

    # ---- Sheet 3: 结算单元汇总 ----
    ws3 = log_wb.create_sheet("结算单元汇总")
    summary_fields = _load_summary_config()
    active_display = []
    active_canonical = []
    for field in summary_fields:
        if field in col_name_idx:
            active_display.append(field)
            active_canonical.append(field)
    header = ["结算单元", *active_display, "人数", "转账总金额"]
    ws3.append(header)
    # 按结算单元分组汇总（工资表侧字段）
    unit_rows = defaultdict(list)
    for out_row, rec in enumerate(payroll_row_map, 1):
        data_rec = all_group[out_row - 1]
        _, _, nrow, _, _, _ = data_rec
        su = str(nrow[col_name_idx["结算单元"]]) if "结算单元" in col_name_idx and col_name_idx["结算单元"] < len(nrow) else "（空）"
        unit_rows[su].append(nrow)
    # 按结算单元统计银行报盘金额
    unit_bank_total = defaultdict(float)
    for pp in payroll_provenance:
        *_, settle_unit = pp
        name = pp[3].strip()
        for bi in bank_by_name.get(name, []):
            try:
                unit_bank_total[settle_unit] += float(bank_prov[bi][4] or 0)
            except (ValueError, TypeError):
                pass
    for settle_unit in sorted(unit_rows, key=str):
        rows = unit_rows[settle_unit]
        row_data = [settle_unit]
        for col in active_canonical:
            idx = col_name_idx[col]
            total = 0.0
            for nrow in rows:
                if idx < len(nrow):
                    try:
                        total += float(nrow[idx] or 0)
                    except (ValueError, TypeError):
                        pass
            row_data.append(round(total, 2))
        row_data.append(len(rows))
        row_data.append(round(unit_bank_total.get(settle_unit, 0.0), 2))
        ws3.append(row_data)

    # 列宽自适应
    for ws in [ws1, ws2, ws3]:
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            col_letter = openpyxl.utils.get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

    log_wb.save(op_log_path)

    return payroll_path, bank_path, validation_results, op_log_path


# ── 标题选择对话框 ──────────────────────────────────

def _choose_title_dialog(root, title_options, subtitle=""):
    """
    弹窗让用户选择输出标题。
    title_options: [(display_text, full_title), ...]
    return: 选中的 full_title，或 None（取消）
    """
    dlg = tk.Toplevel(root)
    dlg.title("选择输出标题")
    dlg.transient(root)
    dlg.grab_set()
    dlg.resizable(False, False)

    w, h = 640, 400
    sw = dlg.winfo_screenwidth()
    sh = dlg.winfo_screenheight()
    dlg.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    tk.Label(dlg, text="检测到多个不同的工资表标题，请选择输出标题：",
             font=("微软雅黑", 11), wraplength=600, justify="left").pack(pady=(15, 5))
    if subtitle:
        tk.Label(dlg, text=subtitle, font=("微软雅黑", 9), fg="gray",
                 wraplength=600, justify="left").pack(pady=(0, 10))

    result = [None]

    frame = tk.Frame(dlg)
    frame.pack(fill="both", expand=True, padx=15, pady=5)

    canvas = tk.Canvas(frame, highlightthickness=0)
    scrollbar = tk.Scrollbar(frame, orient="vertical", command=canvas.yview)
    scrollable = tk.Frame(canvas)

    scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=scrollable, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    var = tk.StringVar(value="")

    # 按频次降序排列
    from collections import Counter
    counts = Counter(t[1] for t in title_options)
    sorted_opts = sorted(set((t[1], t[0]) for t in title_options),
                         key=lambda x: (-counts[x[0]], title_options.index((x[1], x[0]))))

    radiobuttons = []
    for full_title, display_text in sorted_opts:
        cnt = counts.get(full_title, 0)
        label = f"【{cnt} 个文件】{full_title}" if cnt > 1 else full_title
        rb = tk.Radiobutton(scrollable, text=label, variable=var,
                            value=full_title, wraplength=560, anchor="w",
                            justify="left", font=("微软雅黑", 10))
        rb.pack(fill="x", padx=10, pady=3)
        radiobuttons.append(rb)

    if radiobuttons:
        radiobuttons[0].select()

    scrollbar.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)

    btn_frame = tk.Frame(dlg)
    btn_frame.pack(fill="x", padx=15, pady=(5, 15))

    def on_ok():
        result[0] = var.get()
        dlg.destroy()

    def on_cancel():
        dlg.destroy()

    tk.Button(btn_frame, text="确定", width=10, command=on_ok,
              font=("微软雅黑", 10)).pack(side="left", padx=5)
    tk.Button(btn_frame, text="取消", width=10, command=on_cancel,
              font=("微软雅黑", 10)).pack(side="right", padx=5)

    dlg.update()
    canvas.configure(scrollregion=canvas.bbox("all"))
    dlg.wait_window()

    return result[0]


def scan_file_titles(payroll_dir):
    """
    快速扫描 signed_ 工资表，返回所有文件标题行原文。
    return: list of (display_label, full_title_string)
    """
    if not os.path.isdir(payroll_dir):
        return []

    titles = []
    for fname in sorted(os.listdir(payroll_dir)):
        if not fname.startswith("signed_") or not fname.endswith(".xlsx") or fname.startswith("~$"):
            continue
        path = os.path.join(payroll_dir, fname)
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            for cell_val in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
                s = str(cell_val or "").strip()
                if s and s not in ("None",):
                    titles.append((fname, s))
                    break
            wb.close()
        except Exception:
            pass

    return titles
# ──────────────────────────────────────────────


class BatchPrintGUI:
    """批量银行报盘合并与工资表打印工具 GUI"""

    def __init__(self, root):
        self.root = root
        self.root.title("批量打印程序")
        self.root.geometry("800x600")
        self.root.minsize(700, 500)

        self.bank_dir = None
        self.payroll_dir = None
        self.output_dir = None

        self._build_ui()

    # ── UI 构建 ──────────────────────────────

    def _build_ui(self):
        # 标题
        title = tk.Label(
            self.root,
            text="批量银行报盘合并与工资表打印工具",
            font=("微软雅黑", 14, "bold"),
            pady=12,
        )
        title.pack(fill=tk.X)

        # ── 目录选择区域 ──
        dir_frame = tk.Frame(self.root, padx=12, pady=6)
        dir_frame.pack(fill=tk.X)

        tk.Button(dir_frame, text="选择银行报盘目录", command=self.select_bank_dir, width=18).grid(
            row=0, column=0, padx=(0, 8), pady=4
        )
        tk.Button(dir_frame, text="选择工资表目录", command=self.select_payroll_dir, width=18).grid(
            row=0, column=1, padx=8, pady=4
        )
        tk.Button(dir_frame, text="选择输出目录", command=self.select_output_dir, width=18).grid(
            row=0, column=2, padx=8, pady=4
        )

        # 路径显示
        self.bank_dir_label = tk.Label(dir_frame, text="银行报盘目录：未选择", anchor=tk.W, fg="#555")
        self.bank_dir_label.grid(row=1, column=0, columnspan=3, sticky=tk.EW, pady=2)

        self.payroll_dir_label = tk.Label(dir_frame, text="工资表目录：未选择", anchor=tk.W, fg="#555")
        self.payroll_dir_label.grid(row=2, column=0, columnspan=3, sticky=tk.EW, pady=2)

        self.output_dir_label = tk.Label(dir_frame, text="输出目录：未选择", anchor=tk.W, fg="#555")
        self.output_dir_label.grid(row=3, column=0, columnspan=3, sticky=tk.EW, pady=2)

        dir_frame.columnconfigure(0, weight=1)
        dir_frame.columnconfigure(1, weight=1)
        dir_frame.columnconfigure(2, weight=1)

        # ── 按钮区域（目录选择和日志之间） ──
        btn_box = tk.Frame(self.root, padx=12, pady=10)
        btn_box.pack(fill=tk.X)

        # 第一行：主要操作 + 验证规则
        row1 = tk.Frame(btn_box)
        row1.pack(fill=tk.X, pady=(0, 4))

        self.run_all_btn = tk.Button(
            row1,
            text="执行全部并打印",
            command=self.run_all_and_print,
            bg="#4a90d9",
            fg="white",
            font=("微软雅黑", 11, "bold"),
            padx=20,
            pady=4,
        )
        self.run_all_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.merge_only_btn = tk.Button(
            row1,
            text="仅合并不打印",
            command=self.run_merge_only,
            bg="#6abf69",
            fg="white",
            font=("微软雅黑", 11, "bold"),
            padx=20,
            pady=4,
        )
        self.merge_only_btn.pack(side=tk.LEFT)

        # 分隔 + 验证规则
        sep_left = tk.Frame(row1, width=2, bd=1, relief=tk.SUNKEN, height=30)
        sep_left.pack(side=tk.LEFT, padx=10)

        tk.Button(
            row1,
            text="验证规则",
            command=self._open_validation_config,
            bg="#95a5a6",
            fg="white",
            font=("微软雅黑", 10, "bold"),
            padx=12,
            pady=4,
        ).pack(side=tk.LEFT)

        tk.Button(
            row1,
            text="汇总设置",
            command=self._open_summary_config,
            bg="#95a5a6",
            fg="white",
            font=("微软雅黑", 10, "bold"),
            padx=12,
            pady=4,
        ).pack(side=tk.LEFT, padx=(6, 0))

        # 第二行：工资表合并操作
        row2 = tk.Frame(btn_box)
        row2.pack(fill=tk.X)

        self.merge_payroll_btn = tk.Button(
            row2,
            text="合并工资表及报盘并打印",
            command=self.run_merge_payroll_and_print,
            bg="#e67e22",
            fg="white",
            font=("微软雅黑", 11, "bold"),
            padx=16,
            pady=4,
        )
        self.merge_payroll_btn.pack(side=tk.LEFT, padx=(0, 6))

        self.merge_payroll_no_print_btn = tk.Button(
            row2,
            text="合并工资表及报盘不打印",
            command=self.run_merge_payroll_only,
            bg="#9b59b6",
            fg="white",
            font=("微软雅黑", 11, "bold"),
            padx=16,
            pady=4,
        )
        self.merge_payroll_no_print_btn.pack(side=tk.LEFT)

        # 第三行：报盘格式转换
        row3 = tk.Frame(btn_box)
        row3.pack(fill=tk.X, pady=(4, 0))

        self.bank_convert_btn = tk.Button(
            row3,
            text="报盘格式转换",
            command=self.run_bank_convert,
            bg="#8e44ad",
            fg="white",
            font=("微软雅黑", 11, "bold"),
            padx=16,
            pady=4,
        )
        self.bank_convert_btn.pack(side=tk.LEFT, padx=(0, 6))

        self.adv_merge_btn = tk.Button(
            row3,
            text="高级合并报盘",
            command=self.run_advanced_merge,
            bg="#2c3e50",
            fg="white",
            font=("微软雅黑", 11, "bold"),
            padx=16,
            pady=4,
        )
        self.adv_merge_btn.pack(side=tk.LEFT, padx=(6, 0))

        # 分隔线
        sep = tk.Frame(self.root, height=2, bd=1, relief=tk.SUNKEN)
        sep.pack(fill=tk.X, padx=10, pady=6)

        # ── 日志区域 ──
        log_frame = tk.Frame(self.root, padx=12)
        log_frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(log_frame, text="日志信息", font=("微软雅黑", 10, "bold"), anchor=tk.W).pack(
            fill=tk.X, pady=(0, 4)
        )

        self.log_area = scrolledtext.ScrolledText(
            log_frame,
            wrap=tk.WORD,
            font=("Consolas", 10),
            bg="#fafafa",
            state=tk.DISABLED,
        )
        self.log_area.pack(fill=tk.BOTH, expand=True)

    # ── 日志 ──────────────────────────────

    def log(self, message):
        """向日志区域追加一行"""
        self.log_area.config(state=tk.NORMAL)
        self.log_area.insert(tk.END, message + "\n")
        self.log_area.see(tk.END)
        self.log_area.config(state=tk.DISABLED)
        self.root.update_idletasks()

    # ── 目录选择 ──────────────────────────────

    def _set_default_output_dir(self):
        """根据银行报盘目录自动设置输出目录（同级目录）"""
        if not self.bank_dir:
            return
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        parent = os.path.dirname(self.bank_dir)
        default_dir = os.path.join(parent, f"合并后的银行报盘_{ts}")
        self.output_dir = default_dir
        self.output_dir_label.config(text=f"输出目录：{default_dir}")

    def select_bank_dir(self):
        d = filedialog.askdirectory(title="选择银行报盘目录")
        if d:
            self.bank_dir = os.path.normpath(d)
            self.bank_dir_label.config(text=f"银行报盘目录：{os.path.normpath(d)}")
            self._set_default_output_dir()

    def select_payroll_dir(self):
        d = filedialog.askdirectory(title="选择工资表目录")
        if d:
            self.payroll_dir = os.path.normpath(d)
            self.payroll_dir_label.config(text=f"工资表目录：{os.path.normpath(d)}")

    def select_output_dir(self):
        d = filedialog.askdirectory(title="选择输出目录")
        if d:
            self.output_dir = os.path.normpath(d)
            self.output_dir_label.config(text=f"输出目录：{os.path.normpath(d)}")

    # ── 进度控制 ──────────────────────────────

    def _set_busy(self, busy):
        """切换按钮启用状态"""
        state = tk.DISABLED if busy else tk.NORMAL
        self.run_all_btn.config(state=state)
        self.merge_only_btn.config(state=state)
        self.merge_payroll_btn.config(state=state)
        self.merge_payroll_no_print_btn.config(state=state)
        self.root.update_idletasks()

    # ── 执行全部并打印 ──────────────────────────────

    def run_all_and_print(self):
        if not self._check_dirs():
            return

        self._set_busy(True)
        self.log("=" * 50)
        self.log("开始执行全部流程...")
        self.log("")

        # 步骤 1：改名银行报盘文件
        self.log("【步骤1】改名银行报盘文件...")
        try:
            renamed = rename_bank_files(self.bank_dir, self.output_dir)
            self.log(f"  ✓ 改名完成：{len(renamed)} 个文件")
        except Exception as e:
            self.log(f"  ✗ 改名失败：{e}")
            self._set_busy(False)
            return

        # 步骤 2：合并银行报盘文件
        self.log("【步骤2】合并银行报盘文件...")
        try:
            merged, merge_warnings = merge_bank_files(renamed, self.bank_dir, self.output_dir)
            self.log(f"  ✓ 合并完成：{len(merged)} 个合并文件")
            if merge_warnings:
                self.log(f"  ⚠ 金额警告（分以下数值已舍去）：")
                for w in merge_warnings:
                    self.log(f"    - {w}")
        except Exception as e:
            self.log(f"  ✗ 合并失败：{e}")
            self._set_busy(False)
            return

        # 步骤 3：匹配工资表
        self.log("【步骤3】匹配工资表文件...")
        try:
            matched, unmatched, duplicates, no_signed = match_payroll_files(merged, self.payroll_dir)
            matched_count = sum(1 for _, fp, _ in matched if fp is not None)
            self.log(f"  ✓ 匹配完成：{matched_count}/{len(matched)} 匹配成功")
            for merged_name, payroll_path, unit_name in matched:
                status = f"→ {payroll_path}" if payroll_path else "✗ 未签字"
                self.log(f"    {unit_name}: {status}")
            if no_signed:
                self.log(f"  ⚠ 以下 {len(no_signed)} 个单位工资表未签字（无 signed_ 前缀），不能作为匹配依据：")
                for uname, fn in no_signed:
                    self.log(f"    {uname}: {fn}")
            if duplicates:
                self.log(f"  ⚠ 以下单位存在多个工资表文件，仅使用第一个：")
                for unit_name, files in duplicates.items():
                    self.log(f"    {unit_name}:")
                    for f in files:
                        self.log(f"      - {f}")
            if unmatched:
                self.log(f"  ⚠ 以下 {len(unmatched)} 个工资表未匹配到银行报盘，不参与打印：")
                for fname in unmatched:
                    self.log(f"    - {fname}")
        except Exception as e:
            self.log(f"  ✗ 匹配失败：{e}")
            self._set_busy(False)
            return

        # 步骤 4：打印（需要用户确认）
        # 文件名已带序号前缀，字典序即正确顺序
        matched.sort(key=lambda x: x[0])
        self.log("")
        self.log("【步骤4】准备打印...")

        if not check_wps_available():
            self.log("  ✗ WPS Office 不可用，无法打印")
            self.log("  提示：打印功能需要 Windows 环境 + WPS Office")
            self._set_busy(False)
            return

        # 展示合并文件列表供用户确认
        self.log("")
        self.log("以下文件将打印：")
        for merged_name, payroll_path, unit_name in matched:
            if payroll_path:
                self.log(f"  • {unit_name} → {os.path.basename(payroll_path)}")

        self.log("")
        self.log("请在弹出对话框确认后开始打印...")
        self.root.update()

        # 弹出确认对话框
        ok = messagebox.askyesno(
            "确认打印",
            f"将打印 {matched_count} 个单位的工资表，是否继续？",
        )
        if not ok:
            self.log("  用户取消打印")
            self._set_busy(False)
            return

        # 执行打印
        self.log("")
        self.log("开始打印...")

        def progress_cb(current, total, message):
            self.log(f"  [{current}/{total}] {message}")
            self.root.update()

        # 仅打印匹配到工资表的单位
        matched_to_print = [(m, p, u) for m, p, u in matched if p is not None]
        self.log(f"  实际需打印：{len(matched_to_print)} 个（未匹配的 {len(matched) - len(matched_to_print)} 个已跳过）")
        success, fail, fail_list = batch_print(matched_to_print, progress_cb)

        self.log("")
        self.log(f"打印完成：成功 {success}，失败 {fail}")
        if fail_list:
            self.log("失败列表：")
            for merged_name, unit_name in fail_list:
                self.log(f"  ✗ {unit_name} ({merged_name})")

        # 生成操作记录
        try:
            report_path = generate_report_xlsx(
                self.output_dir, renamed, matched,
                unmatched, duplicates, merge_warnings, success, fail, fail_list,
                no_signed=no_signed
            )
            self.log(f"  📄 操作记录已保存：{report_path}")
        except Exception as e:
            self.log(f"  ✗ 操作记录生成失败：{e}")

        self.log("=" * 50)
        self._set_busy(False)

    # ── 仅合并不打印 ──────────────────────────────

    def run_merge_only(self):
        if not self._check_dirs():
            return

        self._set_busy(True)
        self.log("=" * 50)
        self.log("开始执行合并流程（不打印）...")
        self.log("")

        # 步骤 1：改名
        self.log("【步骤1】改名银行报盘文件...")
        try:
            renamed = rename_bank_files(self.bank_dir, self.output_dir)
            self.log(f"  ✓ 改名完成：{len(renamed)} 个文件")
        except Exception as e:
            self.log(f"  ✗ 改名失败：{e}")
            self._set_busy(False)
            return

        # 步骤 2：合并
        self.log("【步骤2】合并银行报盘文件...")
        try:
            merged, merge_warnings = merge_bank_files(renamed, self.bank_dir, self.output_dir)
            self.log(f"  ✓ 合并完成：{len(merged)} 个合并文件")
            if merge_warnings:
                self.log(f"  ⚠ 金额警告（分以下数值已舍去）：")
                for w in merge_warnings:
                    self.log(f"    - {w}")
        except Exception as e:
            self.log(f"  ✗ 合并失败：{e}")
            self._set_busy(False)
            return

        # 步骤 3：匹配
        self.log("【步骤3】匹配工资表文件...")
        try:
            matched, unmatched, duplicates, no_signed = match_payroll_files(merged, self.payroll_dir)
            # 文件名已带序号前缀，字典序即正确顺序
            matched.sort(key=lambda x: x[0])
            matched_count = sum(1 for _, fp, _ in matched if fp is not None)
            self.log(f"  ✓ 匹配完成：{matched_count}/{len(matched)} 匹配成功")
            for merged_name, payroll_path, unit_name in matched:
                status = f"→ {payroll_path}" if payroll_path else "✗ 未签字"
                self.log(f"    {unit_name}: {status}")
            if no_signed:
                self.log(f"  ⚠ 以下 {len(no_signed)} 个单位工资表未签字（无 signed_ 前缀），不能作为匹配依据：")
                for uname, fn in no_signed:
                    self.log(f"    {uname}: {fn}")
            if duplicates:
                self.log(f"  ⚠ 以下单位存在多个工资表文件，仅使用第一个：")
                for unit_name, files in duplicates.items():
                    self.log(f"    {unit_name}:")
                    for f in files:
                        self.log(f"      - {f}")
            if unmatched:
                self.log(f"  ⚠ 以下 {len(unmatched)} 个工资表未匹配到银行报盘，不参与打印：")
                for fname in unmatched:
                    self.log(f"    - {fname}")
        except Exception as e:
            self.log(f"  ✗ 匹配失败：{e}")
            self._set_busy(False)
            return

        self.log("")
        self.log("合并流程完成，未执行打印。")

        # 生成操作记录
        try:
            report_path = generate_report_xlsx(
                self.output_dir, renamed, matched, unmatched, duplicates, merge_warnings,
                no_signed=no_signed
            )
            self.log(f"  📄 操作记录已保存：{report_path}")
        except Exception as e:
            self.log(f"  ✗ 操作记录生成失败：{e}")

        self.log("=" * 50)
        self._set_busy(False)

    # ── 合并工资表及报盘并打印 ──────────────────────────────

    MAX_DIALOG_TITLES = 15  # 超过此数量不弹窗，直接自动构造

    def _prepare_merge(self):
        """返回 (selected_title, cancelled) — cancelled=True 表示用户取消"""
        titles = scan_file_titles(self.payroll_dir)
        unique = sorted(set(t[1] for t in titles))
        if len(unique) <= 1:
            return (unique[0] if unique else None, False)
        if len(unique) > self.MAX_DIALOG_TITLES:
            return (None, False)  # 太多不同标题，自动构造
        chosen = _choose_title_dialog(
            self.root, titles,
            subtitle=f"共 {len(titles)} 个 signed_ 文件，{len(unique)} 种不同标题"
        )
        if chosen is None:
            return (None, True)  # 用户取消
        return (chosen, False)

    def run_merge_payroll_and_print(self):
        if not self._check_dirs():
            return

        self._set_busy(True)
        self.log("=" * 50)
        self.log("开始合并工资表及报盘（按个税分组）...")
        self.log("")

        selected_title, cancelled = self._prepare_merge()
        if cancelled:
            self.log("  用户取消合并")
            self._set_busy(False)
            return
        self.log(f"  输出标题：{selected_title or '（自动合成）'}")

        try:
            payroll_path, bank_path, validation_results, op_log_path = merge_payrolls_by_tax(
                self.payroll_dir, self.output_dir, self.bank_dir,
                selected_title=selected_title
            )
        except Exception as e:
            self.log(f"  ✗ 合并失败：{e}")
            import traceback
            self.log(traceback.format_exc())
            self._set_busy(False)
            return

        self.log(f"  📁 输出目录：{self.output_dir}")
        if payroll_path:
            self.log(f"  ✓ 工资表：{os.path.basename(payroll_path)}")
        if bank_path:
            self.log(f"  ✓ 银行报盘：{os.path.basename(bank_path)}")
        if op_log_path:
            self.log(f"  ✓ 操作记录：{os.path.basename(op_log_path)}")

        # ── 验证结果 ──
        _log_validation_results(self, validation_results)

        # 打印
        self.log("")
        self.log("准备打印...")
        if not check_wps_available():
            self.log("  ✗ WPS Office 不可用，无法打印")
            self._set_busy(False)
            return

        to_print = []
        if payroll_path:
            to_print.append(("", payroll_path, "工资表"))

        self.log("以下文件将打印：")
        for _, fp, label in to_print:
            self.log(f"  • {label} → {os.path.basename(fp)}")

        ok = messagebox.askyesno("确认打印", f"将打印 {len(to_print)} 张工资表，是否继续？")
        if not ok:
            self.log("  用户取消打印")
            self._set_busy(False)
            return

        def progress_cb(current, total, message):
            self.log(f"  [{current}/{total}] {message}")
            self.root.update()

        success, fail, fail_list = batch_print(to_print, progress_cb)
        self.log(f"打印完成：成功 {success}，失败 {fail}")

        self.log("=" * 50)
        self._set_busy(False)

    # ── 合并工资表及报盘不打印 ──────────────────────────────

    def run_merge_payroll_only(self):
        if not self._check_dirs():
            return

        self._set_busy(True)
        self.log("=" * 50)
        self.log("开始合并工资表及报盘（不打印）...")
        self.log("")

        selected_title, cancelled = self._prepare_merge()
        if cancelled:
            self.log("  用户取消合并")
            self._set_busy(False)
            return
        self.log(f"  输出标题：{selected_title or '（自动合成）'}")

        try:
            payroll_path, bank_path, validation_results, op_log_path = merge_payrolls_by_tax(
                self.payroll_dir, self.output_dir, self.bank_dir,
                selected_title=selected_title
            )
        except Exception as e:
            self.log(f"  ✗ 合并失败：{e}")
            import traceback
            self.log(traceback.format_exc())
            self._set_busy(False)
            return

        self.log(f"  📁 输出目录：{self.output_dir}")
        if payroll_path:
            self.log(f"  ✓ 工资表：{os.path.basename(payroll_path)}")
        if bank_path:
            self.log(f"  ✓ 银行报盘：{os.path.basename(bank_path)}")
        if op_log_path:
            self.log(f"  ✓ 操作记录：{os.path.basename(op_log_path)}")

        # ── 验证结果 ──
        _log_validation_results(self, validation_results)

        self.log("")
        self.log("合并完成，未执行打印。")
        self.log("=" * 50)
        self._set_busy(False)

    # ── 报盘格式转换 ──────────────────────────────

    def run_bank_convert(self):
        if not self.bank_dir:
            messagebox.showwarning("提示", "请先选择银行报盘目录")
            return
        if not self.output_dir:
            messagebox.showwarning("提示", "请先选择输出目录")
            return

        self._set_busy(True)
        self.log("=" * 50)
        self.log("开始报盘格式转换...")
        self.log("")

        try:
            out_path, total = convert_bank_format(self.bank_dir, self.output_dir)
        except Exception as e:
            self.log(f"  ✗ 转换失败：{e}")
            import traceback
            self.log(traceback.format_exc())
            self._set_busy(False)
            return

        if out_path:
            self.log(f"  📁 输出目录：{self.output_dir}")
            self.log(f"  ✓ 合并报盘：{os.path.basename(out_path)}（共 {total} 条）")
        else:
            self.log("  ⚠ 未找到可转换的报盘文件")

        self.log("")
        self.log("转换完成。")
        self.log("=" * 50)
        self._set_busy(False)

    # ── 高级合并报盘 ──────────────────────────

    def run_advanced_merge(self):
        if not self.bank_dir:
            messagebox.showwarning("提示", "请先选择银行报盘目录")
            return
        if not self.output_dir:
            messagebox.showwarning("提示", "请先选择输出目录")
            return

        dlg = MergeOptionsDialog(self.root)
        if dlg.cancelled:
            return
        opts = dlg.get_options()

        self._set_busy(True)
        self.log("=" * 50)
        self.log("开始高级合并报盘...")
        self.log("")
        self.log(f"  银行报盘目录：{self.bank_dir}")
        self.log(f"  输出目录：{self.output_dir}")
        self.log(f"  合并选项：")
        self.log(f"    - 同单位不同银行合并：{'✓' if opts['merge_banks'] else '✗'}")
        self.log(f"    - 同单位不同月份合并：{'✓' if opts['merge_months'] else '✗'}")
        self.log(f"    - 按大单位合并：{'✓' if opts['merge_big_org'] else '✗'}")
        self.log(f"    - 文件名格式：{'简洁（大单位名+月份）' if opts['filename_simple'] else '详细（含序号+银行笔数+总金额）'}")

        def progress_cb(current, total, message):
            self.log(f"  [{current}/{total}] {message}")
            self.root.update()

        try:
            output_files, warnings, stats = merge_bank_files_advanced(
                self.bank_dir, self.output_dir,
                merge_different_banks=opts["merge_banks"],
                merge_different_months=opts["merge_months"],
                merge_by_big_org=opts["merge_big_org"],
                filename_simple=opts["filename_simple"],
                progress_callback=progress_cb,
            )
        except Exception as e:
            self.log(f"  ✗ 合并失败：{e}")
            import traceback
            self.log(traceback.format_exc())
            self._set_busy(False)
            return

        self.log("")
        self.log(f"  📁 输出目录：{self.output_dir}")
        self.log(f"  ✓ 生成 {len(output_files)} 个合并文件：")
        for fpath in sorted(output_files):
            self.log(f"    {os.path.basename(fpath)}")

        if warnings:
            self.log(f"  ⚠ 警告 ({len(warnings)} 条)：")
            for w in warnings:
                self.log(f"    - {w}")

        self.log("")
        self.log("  📊 统计信息：")
        self.log(f"    原文件数：{stats['total_files']}")
        self.log(f"    合并文件数：{stats['output_count']}")
        self.log(f"    大单位数：{len(stats['big_orgs'])}")
        excluded_units = stats.get("excluded_units", set())
        if excluded_units:
            self.log(f"    排除项单位数：{len(excluded_units)}")
            for eu in sorted(excluded_units):
                self.log(f"      - {eu}")

        self.log("")
        self.log("高级合并完成。")
        self.log("=" * 50)
        self._set_busy(False)

    # ── 验证规则配置 ──────────────────────────

    def _open_validation_config(self):
        """打开验证规则配置对话框"""
        ValidationConfigDialog(self.root, self.log)

    # ── 汇总设置 ─────────────────────────────

    def _open_summary_config(self):
        """打开结算单元汇总设置对话框"""
        SummaryConfigDialog(self.root, self.log)


    # ── 目录检查 ──────────────────────────────

    def _check_dirs(self):
        """检查三个目录是否都已选择"""
        if not self.bank_dir:
            messagebox.showwarning("提示", "请先选择银行报盘目录")
            return False
        if not self.payroll_dir:
            messagebox.showwarning("提示", "请先选择工资表目录")
            return False
        if not self.output_dir:
            messagebox.showwarning("提示", "请先选择输出目录")
            return False
        return True


class ValidationConfigDialog:
    """验证规则配置对话框"""

    def __init__(self, parent, log_callback=None):
        self.parent = parent
        self.log_callback = log_callback
        self.cfg = _load_validation_config()

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("验证规则配置")
        self.dialog.transient(parent)
        self.dialog.resizable(True, True)

        main_frame = tk.Frame(self.dialog, padx=16, pady=12)
        main_frame.pack(fill=tk.BOTH, expand=True)

        row = 0

        # ── 容差 ──
        tol_frame = tk.Frame(main_frame)
        tol_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Label(tol_frame, text="容差:", width=8, anchor=tk.W).pack(side=tk.LEFT)
        self.tol_var = tk.StringVar(value=str(self.cfg.get("tolerance", 0.005)))
        tk.Spinbox(tol_frame, from_=0, to=0.1, increment=0.001,
                   textvariable=self.tol_var, width=10).pack(side=tk.LEFT)

        # ── 行公式 ──
        sep = tk.Frame(main_frame, height=1, bd=1, relief=tk.SUNKEN)
        sep.pack(fill=tk.X, pady=4)
        tk.Label(main_frame, text="行公式（LHS = RHS 之和）",
                 font=("微软雅黑", 10, "bold"), anchor=tk.W).pack(fill=tk.X, pady=(4, 4))

        formulas_frame = tk.Frame(main_frame)
        formulas_frame.pack(fill=tk.BOTH, expand=True)

        # Canvas + Scrollbar for formulas
        canvas = tk.Canvas(formulas_frame, highlightthickness=0)
        scrollbar = tk.Scrollbar(formulas_frame, orient=tk.VERTICAL, command=canvas.yview)
        self.formulas_inner = tk.Frame(canvas)

        self.formulas_inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.formulas_inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 绑定鼠标滚轮（绑定到 dialog，关闭时自动清理）
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self.dialog.bind("<MouseWheel>", _on_mousewheel)
        self.dialog.protocol("WM_DELETE_WINDOW", self._on_close)

        self.formula_widgets = []
        self._rebuild_formula_list()

        # ── 公式按钮行 ──
        btn_row = tk.Frame(main_frame)
        btn_row.pack(fill=tk.X, pady=4)
        tk.Button(btn_row, text="编辑选中公式", command=self._edit_selected_formula,
                  width=14).pack(side=tk.LEFT, padx=(0, 6))
        tk.Button(btn_row, text="重置为默认", command=self._reset_formulas,
                  width=14).pack(side=tk.LEFT)

        # ── 列加总校验 & 关键列 ──
        sep2 = tk.Frame(main_frame, height=1, bd=1, relief=tk.SUNKEN)
        sep2.pack(fill=tk.X, pady=6)

        bottom_frame = tk.Frame(main_frame)
        bottom_frame.pack(fill=tk.X)

        # 左半：列加总
        left_frame = tk.LabelFrame(bottom_frame, text="列加总校验", padx=8, pady=4)
        left_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.colsum_vars = {}
        all_known_cols = [
            "基本工资", "补发工资", "应发工资", "单位缴纳五险一金",
            "单位代理费", "雇主责任险", "大病险", "转账合计",
            "个人所得税", "实发工资", "实发合计",
        ]
        existing_targets = [t[0] for t in self.cfg.get("column_sum_targets", [])]
        for i, col in enumerate(all_known_cols):
            var = tk.BooleanVar(value=col in existing_targets)
            self.colsum_vars[col] = var
            cb = tk.Checkbutton(left_frame, text=col, variable=var)
            cb.grid(row=i // 2, column=i % 2, sticky=tk.W, padx=4)

        # 右半：关键列
        right_frame = tk.LabelFrame(bottom_frame, text="关键列（必须存在）", padx=8, pady=4)
        right_frame.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(8, 0))
        self.crit_vars = {}
        all_crit = ["转账合计", "扣款合计", "实发合计", "应发工资"]
        existing_crit = list(self.cfg.get("critical_columns", {}).keys())
        for i, col in enumerate(all_crit):
            var = tk.BooleanVar(value=col in existing_crit)
            self.crit_vars[col] = var
            tk.Checkbutton(right_frame, text=col, variable=var).grid(
                row=i, column=0, sticky=tk.W, padx=4)

        # ── 底部按钮 ──
        btn_bar = tk.Frame(main_frame)
        btn_bar.pack(fill=tk.X, pady=(10, 0))

        tk.Button(btn_bar, text="保存", command=self._save_config,
                  bg="#4a90d9", fg="white", font=("微软雅黑", 10, "bold"),
                  padx=20, pady=2).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn_bar, text="恢复默认", command=self._restore_default,
                  padx=12, pady=2).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn_bar, text="关闭", command=self._on_close,
                  padx=12, pady=2).pack(side=tk.RIGHT)

        # ── 布局完成后自适应大小并居中 ──
        self.dialog.update_idletasks()
        self.dialog.minsize(780, 560)
        # 如果自然尺寸小于最小值，用最小值
        w = max(self.dialog.winfo_reqwidth(), 780)
        h = max(self.dialog.winfo_reqheight(), 560)
        # 居中于父窗口
        pw = parent.winfo_width()
        ph = parent.winfo_height()
        px = parent.winfo_rootx()
        py = parent.winfo_rooty()
        x = px + (pw - w) // 2
        y = py + (ph - h) // 2
        self.dialog.geometry(f"{w}x{h}+{max(0,x)}+{max(0,y)}")
        self.dialog.grab_set()

        self.dialog.wait_window()

    def _on_close(self):
        self.dialog.destroy()

    def _rebuild_formula_list(self):
        for w in self.formula_widgets:
            w.destroy()
        self.formula_widgets = []
        self.selected_formula_idx = tk.IntVar(value=0)

        formulas = self.cfg.get("row_formulas", [])
        if not formulas:
            lbl = tk.Label(self.formulas_inner, text="（无公式）", fg="#999")
            lbl.pack(anchor=tk.W, padx=4, pady=2)
            self.formula_widgets.append(lbl)
            return

        for i, f in enumerate(formulas):
            frame = tk.Frame(self.formulas_inner)
            frame.pack(fill=tk.X, pady=1)

            name = f.get("name", f"公式{i+1}")
            if f.get("rhs_subtract"):
                summary = f"{name}"
            else:
                rhs_list = f.get("rhs", [])
                rhs_parts = [kw[0] if kw else "?" for kw in rhs_list]
                summary = f"{name}"

            rb = tk.Radiobutton(frame, variable=self.selected_formula_idx,
                                value=i, anchor=tk.W)
            rb.pack(side=tk.LEFT)

            lbl = tk.Label(frame, text=summary, anchor=tk.W, font=("微软雅黑", 9))
            lbl.pack(side=tk.LEFT, fill=tk.X, expand=True)

            self.formula_widgets.append(frame)

    def _edit_selected_formula(self):
        idx = self.selected_formula_idx.get()
        formulas = self.cfg.get("row_formulas", [])
        if idx < 0 or idx >= len(formulas):
            return
        _FormulaEditDialog(self.dialog, self.cfg, idx, self._rebuild_formula_list)

    def _reset_formulas(self):
        default = _get_default_validation_config()
        self.cfg["row_formulas"] = default["row_formulas"]
        self._rebuild_formula_list()

    def _restore_default(self):
        default = _get_default_validation_config()
        self.cfg = default
        self.tol_var.set(str(default["tolerance"]))
        # Reset checkboxes
        for col, var in self.colsum_vars.items():
            var.set(col in [t[0] for t in default.get("column_sum_targets", [])])
        for col, var in self.crit_vars.items():
            crit_keys = list(default.get("critical_columns", {}).keys())
            var.set(col in crit_keys)
        self._rebuild_formula_list()

    def _save_config(self):
        # 更新容差
        try:
            tol = float(self.tol_var.get())
            self.cfg["tolerance"] = tol
        except ValueError:
            messagebox.showerror("错误", "容差必须是数字", parent=self.dialog)
            return

        # 更新列加总
        targets = []
        for col, var in self.colsum_vars.items():
            if var.get():
                targets.append([col])
        self.cfg["column_sum_targets"] = targets

        # 更新关键列
        crit = {}
        for col, var in self.crit_vars.items():
            if var.get():
                if col == "扣款合计":
                    crit[col] = ["扣款明细/扣款合计", "扣款合计"]
                else:
                    crit[col] = [col]
        self.cfg["critical_columns"] = crit

        try:
            _save_validation_config(self.cfg)
            if self.log_callback:
                self.log_callback("  ✓ 验证规则配置已保存")
            messagebox.showinfo("成功", "验证规则配置已保存", parent=self.dialog)
        except Exception as e:
            messagebox.showerror("错误", f"保存失败: {e}", parent=self.dialog)


class SummaryConfigDialog:
    """结算单元汇总字段配置对话框"""

    def __init__(self, parent, log_callback=None):
        self.parent = parent
        self.log_callback = log_callback
        self.fields = _load_summary_config()

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("结算单元汇总设置")
        self.dialog.transient(parent)
        self.dialog.resizable(True, True)

        main_frame = tk.Frame(self.dialog, padx=16, pady=12)
        main_frame.pack(fill=tk.BOTH, expand=True)

        tk.Label(main_frame, text="选择需要在结算单元汇总表中显示的字段：",
                 font=("微软雅黑", 10, "bold"), anchor=tk.W).pack(fill=tk.X, pady=(0, 4))
        tk.Label(main_frame, text="勾选预设列或在下方输入自定义列名（工资表中实际存在的列才生效，不存在的自动跳过）",
                 font=("微软雅黑", 9), fg="#666", anchor=tk.W).pack(fill=tk.X, pady=(0, 8))

        # ── 预设列复选框 ──
        preset_frame = tk.LabelFrame(main_frame, text="预设列", padx=8, pady=4)
        preset_frame.pack(fill=tk.X, pady=(0, 8))

        all_known_cols = [
            "基本工资", "补发工资", "应发工资", "单位缴纳五险一金",
            "单位代理费", "雇主责任险", "大病险", "转账合计",
            "个人所得税", "个人工会会费", "工会经费",
            "实发工资", "实发合计", "扣款合计",
        ]
        self.check_vars = {}
        for i, col in enumerate(all_known_cols):
            var = tk.BooleanVar(value=col in self.fields)
            self.check_vars[col] = var
            tk.Checkbutton(preset_frame, text=col, variable=var, anchor=tk.W,
                           font=("微软雅黑", 9)).grid(row=i // 4, column=i % 4, sticky=tk.W, padx=6, pady=1)

        # ── 自定义列 ──
        custom_frame = tk.LabelFrame(main_frame, text="自定义列", padx=8, pady=4)
        custom_frame.pack(fill=tk.X, pady=(0, 8))

        entry_row = tk.Frame(custom_frame)
        entry_row.pack(fill=tk.X, pady=(0, 4))
        self.custom_entry = tk.Entry(entry_row, font=("微软雅黑", 10))
        self.custom_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        self.custom_entry.bind("<Return>", lambda e: self._add_custom())
        tk.Button(entry_row, text="添加", command=self._add_custom,
                  font=("微软雅黑", 9), padx=10).pack(side=tk.RIGHT)

        # 已添加的自定义列标签
        self.custom_tags_frame = tk.Frame(custom_frame)
        self.custom_tags_frame.pack(fill=tk.X)
        self.custom_items = [c for c in self.fields if c not in all_known_cols]
        self._rebuild_custom_tags()

        # ── 底部按钮 ──
        btn_bar = tk.Frame(main_frame)
        btn_bar.pack(fill=tk.X, pady=(10, 0))

        tk.Button(btn_bar, text="保存", command=self._save_config,
                  bg="#4a90d9", fg="white", font=("微软雅黑", 10, "bold"),
                  padx=20, pady=2).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn_bar, text="全选", command=self._select_all,
                  padx=12, pady=2).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn_bar, text="取消全选", command=self._deselect_all,
                  padx=12, pady=2).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn_bar, text="关闭", command=self._on_close,
                  padx=12, pady=2).pack(side=tk.RIGHT)

        # 居中并显示
        self.dialog.update_idletasks()
        w = max(self.dialog.winfo_reqwidth(), 460)
        h = max(self.dialog.winfo_reqheight(), 400)
        pw = parent.winfo_width()
        ph = parent.winfo_height()
        px = parent.winfo_rootx()
        py = parent.winfo_rooty()
        x = px + (pw - w) // 2
        y = py + (ph - h) // 2
        self.dialog.geometry(f"{w}x{h}+{max(0,x)}+{max(0,y)}")
        self.dialog.grab_set()
        self.dialog.wait_window()

    def _add_custom(self):
        name = self.custom_entry.get().strip()
        if not name:
            return
        if name in self.custom_items:
            messagebox.showinfo("提示", f"「{name}」已在列表中", parent=self.dialog)
            return
        self.custom_items.append(name)
        self.custom_entry.delete(0, tk.END)
        self._rebuild_custom_tags()

    def _rebuild_custom_tags(self):
        for w in self.custom_tags_frame.winfo_children():
            w.destroy()
        if not self.custom_items:
            tk.Label(self.custom_tags_frame, text="（无，在输入框中输入列名后点击添加）",
                     fg="#999", font=("微软雅黑", 9)).pack(anchor=tk.W, padx=4, pady=2)
            return
        for item in self.custom_items:
            tag = tk.Frame(self.custom_tags_frame, bd=1, relief=tk.RIDGE, padx=4, pady=1)
            tag.pack(side=tk.LEFT, padx=3, pady=2)
            tk.Label(tag, text=item, font=("微软雅黑", 9)).pack(side=tk.LEFT)
            tk.Label(tag, text=" ×", fg="red", font=("微软雅黑", 9, "bold"),
                     cursor="hand2").pack(side=tk.LEFT, padx=(4, 0))
            # 用 lambda 绑定删除
            tag.bind("<Button-1>", lambda e, x=item: self._remove_custom(x))
            # 让标签内所有子部件也响应点击
            for child in tag.winfo_children():
                child.bind("<Button-1>", lambda e, x=item: self._remove_custom(x))

    def _remove_custom(self, name):
        if name in self.custom_items:
            self.custom_items.remove(name)
            self._rebuild_custom_tags()

    def _select_all(self):
        for var in self.check_vars.values():
            var.set(True)

    def _deselect_all(self):
        for var in self.check_vars.values():
            var.set(False)

    def _save_config(self):
        selected = [col for col, var in self.check_vars.items() if var.get()]
        selected.extend(self.custom_items)
        try:
            _save_summary_config(selected)
            if self.log_callback:
                self.log_callback(f"  ✓ 结算单元汇总设置已保存（{len(selected)} 个字段）")
            messagebox.showinfo("成功", "结算单元汇总设置已保存", parent=self.dialog)
        except Exception as e:
            messagebox.showerror("错误", f"保存失败: {e}", parent=self.dialog)

    def _on_close(self):
        self.dialog.destroy()
        btn_bar.pack(fill=tk.X, pady=(10, 0))

        tk.Button(btn_bar, text="保存", command=self._save_config,
                  bg="#4a90d9", fg="white", font=("微软雅黑", 10, "bold"),
                  padx=20, pady=2).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn_bar, text="全选", command=self._select_all,
                  padx=12, pady=2).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn_bar, text="取消全选", command=self._deselect_all,
                  padx=12, pady=2).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn_bar, text="关闭", command=self._on_close,
                  padx=12, pady=2).pack(side=tk.RIGHT)

        # 居中并显示
        self.dialog.update_idletasks()
        w = max(self.dialog.winfo_reqwidth(), 320)
        h = max(self.dialog.winfo_reqheight(), 420)
        pw = parent.winfo_width()
        ph = parent.winfo_height()
        px = parent.winfo_rootx()
        py = parent.winfo_rooty()
        x = px + (pw - w) // 2
        y = py + (ph - h) // 2
        self.dialog.geometry(f"{w}x{h}+{max(0,x)}+{max(0,y)}")
        self.dialog.grab_set()
        self.dialog.wait_window()

    def _select_all(self):
        for var in self.check_vars.values():
            var.set(True)

    def _deselect_all(self):
        for var in self.check_vars.values():
            var.set(False)

    def _save_config(self):
        selected = [col for col, var in self.check_vars.items() if var.get()]
        try:
            _save_summary_config(selected)
            if self.log_callback:
                self.log_callback(f"  ✓ 结算单元汇总设置已保存（{len(selected)} 个字段）")
            messagebox.showinfo("成功", "结算单元汇总设置已保存", parent=self.dialog)
        except Exception as e:
            messagebox.showerror("错误", f"保存失败: {e}", parent=self.dialog)

    def _on_close(self):
        self.dialog.destroy()


class _FormulaEditDialog:
    """行公式编辑子对话框"""

    def __init__(self, parent, cfg, formula_idx, on_save_callback):
        self.cfg = cfg
        self.formula_idx = formula_idx
        self.on_save_callback = on_save_callback
        formula = cfg["row_formulas"][formula_idx]

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("编辑公式")
        self.dialog.transient(parent)
        self.dialog.resizable(True, True)

        main = tk.Frame(self.dialog, padx=12, pady=10)
        main.pack(fill=tk.BOTH, expand=True)

        # 公式名称
        tk.Label(main, text="公式名称:").pack(anchor=tk.W)
        self.name_var = tk.StringVar(value=formula.get("name", ""))
        tk.Entry(main, textvariable=self.name_var).pack(fill=tk.X, pady=(0, 8))

        # LHS 关键词
        tk.Label(main, text="LHS (左侧列关键词, 逗号分隔):").pack(anchor=tk.W)
        lhs_kw_list = formula.get("lhs", [[""]])
        lhs_str = ",".join(lhs_kw_list[0]) if lhs_kw_list else ""
        self.lhs_var = tk.StringVar(value=lhs_str)
        tk.Entry(main, textvariable=self.lhs_var).pack(fill=tk.X, pady=(0, 8))

        is_subtract = formula.get("rhs_subtract", False)

        if is_subtract:
            tk.Label(main, text="此公式为扣款子项汇总（自动计算所有子项之和）",
                     fg="#666").pack(anchor=tk.W, pady=8)
        else:
            # RHS 列表
            tk.Label(main, text="RHS (右侧列, 每行一个关键词列表):",
                     font=("微软雅黑", 9, "bold")).pack(anchor=tk.W)

            rhs_frame = tk.Frame(main)
            rhs_frame.pack(fill=tk.BOTH, expand=True)

            canvas = tk.Canvas(rhs_frame, highlightthickness=0)
            scrollbar = tk.Scrollbar(rhs_frame, orient=tk.VERTICAL, command=canvas.yview)
            self.rhs_inner = tk.Frame(canvas)
            self.rhs_inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            canvas.create_window((0, 0), window=self.rhs_inner, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            def _mw(event):
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            self.dialog.bind("<MouseWheel>", _mw)

            self.rhs_entries = []
            rhs_list = formula.get("rhs", [])
            if not rhs_list:
                rhs_list = [[""]]
            for kw_list in rhs_list:
                self._add_rhs_row(kw_list)

            btn_row = tk.Frame(main)
            btn_row.pack(fill=tk.X, pady=4)
            tk.Button(btn_row, text="+ 添加列", command=lambda: self._add_rhs_row([""]),
                      width=10).pack(side=tk.LEFT, padx=(0, 6))
            tk.Button(btn_row, text="移除末项", command=self._remove_last_rhs,
                      width=10).pack(side=tk.LEFT)

        # ── 底部按钮 ──
        bar = tk.Frame(main)
        bar.pack(fill=tk.X, pady=(8, 0))
        tk.Button(bar, text="确定", command=self._save,
                  bg="#4a90d9", fg="white", padx=16).pack(side=tk.RIGHT, padx=(4, 0))
        tk.Button(bar, text="取消", command=self.dialog.destroy,
                  padx=12).pack(side=tk.RIGHT)

        self.dialog.protocol("WM_DELETE_WINDOW", self.dialog.destroy)

        # ── 布局完成后自适应大小并居中 ──
        self.dialog.update_idletasks()
        self.dialog.minsize(560, 420)
        w = max(self.dialog.winfo_reqwidth(), 560)
        h = max(self.dialog.winfo_reqheight(), 420)
        pw = parent.winfo_width()
        ph = parent.winfo_height()
        px = parent.winfo_rootx()
        py = parent.winfo_rooty()
        x = px + (pw - w) // 2
        y = py + (ph - h) // 2
        self.dialog.geometry(f"{w}x{h}+{max(0,x)}+{max(0,y)}")
        self.dialog.grab_set()

    def _add_rhs_row(self, kw_list):
        frame = tk.Frame(self.rhs_inner)
        frame.pack(fill=tk.X, pady=1)

        label = tk.Label(frame, text=f"项 {len(self.rhs_entries) + 1}:", width=5, anchor=tk.W)
        label.pack(side=tk.LEFT)

        entry = tk.Entry(frame)
        entry.insert(0, ",".join(kw_list))
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))

        self.rhs_entries.append(entry)

    def _remove_last_rhs(self):
        if self.rhs_entries:
            entry = self.rhs_entries.pop()
            entry.master.destroy()

    def _save(self):
        formula = self.cfg["row_formulas"][self.formula_idx]
        formula["name"] = self.name_var.get().strip()

        # 更新 LHS
        lhs_text = self.lhs_var.get().strip()
        lhs_kws = [kw.strip() for kw in lhs_text.split(",") if kw.strip()]
        if lhs_kws:
            formula["lhs"] = [lhs_kws]
        else:
            formula["lhs"] = [[""]]

        if not formula.get("rhs_subtract"):
            # 更新 RHS
            rhs = []
            for entry in self.rhs_entries:
                text = entry.get().strip()
                if text:
                    kws = [kw.strip() for kw in text.split(",") if kw.strip()]
                    rhs.append(kws if kws else [text])
            formula["rhs"] = rhs

        self.on_save_callback()
        self.dialog.destroy()


class MergeOptionsDialog:
    """高级合并报盘选项对话框"""

    def __init__(self, parent):
        self.cancelled = True
        self._options = {
            "merge_banks": True,
            "merge_months": True,
            "merge_big_org": True,
            "filename_simple": True,
        }

        self.dialog = tk.Toplevel(parent)
        self.dialog.title("高级合并报盘选项")
        self.dialog.transient(parent)
        self.dialog.resizable(False, False)
        self.dialog.grab_set()

        main = tk.Frame(self.dialog, padx=20, pady=16)
        main.pack(fill=tk.BOTH, expand=True)

        tk.Label(main, text="合并选项（默认全部开启）：",
                 font=("微软雅黑", 11, "bold"), anchor=tk.W).pack(fill=tk.X, pady=(0, 12))

        options = [
            ("merge_banks", "合并同单位不同银行", "同一单位的多家银行报盘合并为一个文件"),
            ("merge_months", "合并同单位不同月份", "同一单位的多个月份报盘合并为一个文件"),
            ("merge_big_org", "按大单位合并", "将同属一个大单位的多个子单位合并，排除项除外"),
        ]

        self.vars = {}
        for key, label, desc in options:
            frame = tk.Frame(main)
            frame.pack(fill=tk.X, pady=6)
            var = tk.BooleanVar(value=True)
            self.vars[key] = var
            cb = tk.Checkbutton(frame, text=label, variable=var,
                                font=("微软雅黑", 10, "bold"))
            cb.pack(anchor=tk.W)
            tk.Label(frame, text=desc, font=("微软雅黑", 9), fg="#666",
                     anchor=tk.W).pack(anchor=tk.W, padx=(24, 0))

        tk.Frame(main, height=1, bd=1, relief=tk.SUNKEN).pack(fill=tk.X, pady=10)

        tk.Label(main, text="文件名格式：",
                 font=("微软雅黑", 10, "bold"), anchor=tk.W).pack(fill=tk.X, pady=(0, 6))
        self.filename_var = tk.StringVar(value="simple")
        for fval, flabel, fdesc in [
            ("simple", "简洁（推荐）", "大单位名-月份.xlsx（不含序号）"),
            ("full", "详细", "序号_大单位名-月份-银行笔数-总数-总金额.xlsx"),
        ]:
            frame = tk.Frame(main)
            frame.pack(fill=tk.X, pady=2)
            rb = tk.Radiobutton(frame, text=flabel, variable=self.filename_var,
                                value=fval, font=("微软雅黑", 10))
            rb.pack(anchor=tk.W, side=tk.LEFT)
            tk.Label(frame, text=fdesc, font=("微软雅黑", 9), fg="#888",
                     anchor=tk.W).pack(anchor=tk.W, side=tk.LEFT, padx=(6, 0))

        tk.Frame(main, height=1, bd=1, relief=tk.SUNKEN).pack(fill=tk.X, pady=10)
        info_text = (
            "说明：\n"
            "• 排除项（外包业务等）不参与按大单位合并，独立生成文件\n"
            "• 输出格式：吉林银行代发业务导入模板\n"
            "• 金额为 0 的数据：不写入输出报盘，在操作记录.xlsx 中保留"
        )
        tk.Label(main, text=info_text, font=("微软雅黑", 9), fg="#555",
                 justify=tk.LEFT, anchor=tk.W).pack(fill=tk.X, pady=(0, 12))

        btn_frame = tk.Frame(main)
        btn_frame.pack(fill=tk.X)
        tk.Button(btn_frame, text="开始合并", command=self._on_ok,
                  bg="#4a90d9", fg="white", font=("微软雅黑", 10, "bold"),
                  padx=20, pady=2).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn_frame, text="取消", command=self._on_cancel,
                  padx=12, pady=2).pack(side=tk.RIGHT)

        self.dialog.update_idletasks()
        w = max(self.dialog.winfo_reqwidth(), 420)
        h = max(self.dialog.winfo_reqheight(), 320)
        pw = parent.winfo_width()
        ph = parent.winfo_height()
        px = parent.winfo_rootx()
        py = parent.winfo_rooty()
        x = px + (pw - w) // 2
        y = py + (ph - h) // 2
        self.dialog.geometry(f"{w}x{h}+{max(0,x)}+{max(0,y)}")

        self.dialog.wait_window()

    def _on_ok(self):
        self._options = {
            "merge_banks": self.vars["merge_banks"].get(),
            "merge_months": self.vars["merge_months"].get(),
            "merge_big_org": self.vars["merge_big_org"].get(),
            "filename_simple": self.filename_var.get() == "simple",
        }
        self.cancelled = False
        self.dialog.destroy()

    def _on_cancel(self):
        self.cancelled = True
        self.dialog.destroy()

    def get_options(self):
        return self._options


def main():
    root = tk.Tk()
    app = BatchPrintGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
