"""
batchprint_gui.py 核心逻辑集成测试

测试流程:
  1. 改名测试: 107个银行报盘文件 → 改名到临时目录
  2. 合并列映射测试: 97个单位合并, 工商银行3→9列映射
  3. 匹配测试: 97/97 匹配工资表, signed_优先
  4. 排序测试: 合并文件按文件名升序

用法: python3 test_integration.py
"""

import os
import sys
import shutil
import tempfile
import openpyxl

# 确保能导入 batchprint_gui
sys.path.insert(0, "/home/ubuntu/coding/seqprint")
from batchprint_gui import (
    rename_bank_files,
    merge_bank_files,
    match_payroll_files,
    split_filename,
    detect_bank_type,
    _read_icbc_rows,
    _read_ccb_rows,
    _read_jlb_rows,
)

BANK_DIR = "/home/ubuntu/excel_example/baiyun/银行报盘"
PAYROLL_DIR = "/home/ubuntu/excel_example/baiyun/202607090002_张朦_吉林大学校医院、吉林大学本科生院等97家单位2026年06月工资发放请示"


def log(msg):
    print(msg)


def check(cond, msg):
    if cond:
        log(f"  ✓ {msg}")
    else:
        log(f"  ✗ {msg}")
        global HAS_FAILURE
        HAS_FAILURE = True


HAS_FAILURE = False


# ──────────────────────────────────────────────
# 测试 1: 改名测试
# ──────────────────────────────────────────────
def test_rename():
    log("\n" + "=" * 60)
    log("【测试 1】改名测试")
    log("=" * 60)

    tmpdir = tempfile.mkdtemp(prefix="test_rename_")
    try:
        # 列出所有 .xls
        all_bank_files = [f for f in os.listdir(BANK_DIR) if f.lower().endswith(".xls")]
        check(len(all_bank_files) == 107, f"银行报盘 .xls 文件数量: {len(all_bank_files)} (期望 107)")

        # 执行改名
        renamed = rename_bank_files(BANK_DIR, tmpdir)

        # 验证数量
        check(len(renamed) == 107, f"改名返回数量: {len(renamed)} (期望 107)")

        # 验证输出目录文件数量
        output_files = [f for f in os.listdir(tmpdir) if f.lower().endswith(".xls")]
        check(len(output_files) == 107, f"改名后目录文件数: {len(output_files)} (期望 107)")

        # 验证每个新文件名格式: 单位名-YYYYMM-银行名称.xls
        format_ok = True
        for new_name, old_name, bank_name, unit_name in renamed:
            # 新文件名格式: {unit_name}-{yearmon}-{bank_name}.xls
            parts = new_name.rsplit(".", 1)[0].split("-", 2)
            if len(parts) != 3:
                format_ok = False
                log(f"    格式异常: {new_name}")
                continue
            u, y, b = parts
            if y not in old_name or b not in old_name:
                format_ok = False
                log(f"    内容不匹配: new={new_name}, old={old_name}")
        check(format_ok, "所有新文件名格式为 单位名-YYYYMM-银行名称.xls")

        # 验证带括号的特殊文件
        paren_files = [r for r in renamed if "（" in r[0] or "）" in r[0] or "(" in r[0] or ")" in r[0]]
        check(len(paren_files) >= 1, f"含括号文件名正确处理: {len(paren_files)} 个")

        # 验证 split_filename 反向解析一致
        for new_name, _, _, _ in renamed:
            try:
                y2, b2, u2 = split_filename(new_name)
            except Exception as e:
                log(f"    split_filename 失败: {new_name} → {e}")
                format_ok = False
        check(format_ok, "所有新文件名可被 split_filename 正确解析")

        # 验证唯一单位数
        units_from_rename = {}
        for _, _, _, unit_name in renamed:
            units_from_rename.setdefault(unit_name, 0)
            units_from_rename[unit_name] += 1
        check(len(units_from_rename) == 97, f"唯一单位数: {len(units_from_rename)} (期望 97)")

        # 验证多银行单位
        multi = {u: c for u, c in units_from_rename.items() if c > 1}
        check(len(multi) == 9, f"多银行单位数: {len(multi)} (期望 9)")

        return renamed
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


# ──────────────────────────────────────────────
# 测试 2: 合并列映射测试
# ──────────────────────────────────────────────
def test_merge(renamed_list):
    log("\n" + "=" * 60)
    log("【测试 2】合并列映射测试")
    log("=" * 60)

    tmpdir = tempfile.mkdtemp(prefix="test_merge_")
    try:
        merged = merge_bank_files(renamed_list, BANK_DIR, tmpdir)

        check(len(merged) == 97, f"合并文件数: {len(merged)} (期望 97)")

        # 验证输出文件
        output_files = [f for f in os.listdir(tmpdir) if f.lower().endswith(".xlsx")]
        check(len(output_files) == 97, f"输出目录 .xlsx 文件数: {len(output_files)} (期望 97)")

        # 验证没有 .xls 文件残留在输出目录
        xls_left = [f for f in os.listdir(tmpdir) if f.lower().endswith(".xls")]
        check(len(xls_left) == 0, f"输出目录无 .xls 残留: {len(xls_left)} 个 (期望 0)")

        # 验证合并文件命名格式: 单位名-YYYYMM-银行1行数-银行2行数.xlsx
        naming_ok = True
        for m in merged:
            if not m.endswith(".xlsx"):
                naming_ok = False
                log(f"    非 .xlsx 后缀: {m}")
            parts = m.rsplit(".", 1)[0].split("-")
            if len(parts) < 3:
                naming_ok = False
                log(f"    命名格式异常: {m}")
        check(naming_ok, "所有合并文件为 .xlsx 格式且命名包含单位-年月-银行行数")

        # 验证多银行单位的合并文件名正确
        # 文件名含多个银行段: 单位名-202606-建设银行N-工商银行M[-(更多银行)].xlsx
        multi_merged = [m for m in merged if len(m.split("-")) > 3]  # 2家银行=4段, 3家=5段
        check(len(multi_merged) == 9, f"多银行合并文件数: {len(multi_merged)} (期望 9)")

        # 详细验证工商银行列映射
        log("  ── 工商银行 3→9 列映射验证 ──")
        icbc_units = [r for r in renamed_list if r[2] == "工商银行"]
        check(len(icbc_units) > 0, f"工商银行文件数: {len(icbc_units)}")

        icbc_ok = True
        for new_name, old_name, bank_name, unit_name in icbc_units:
            filepath = os.path.join(BANK_DIR, old_name)
            rows = _read_icbc_rows(filepath)
            for row in rows:
                # [seq, account, name, amount, "1", "工商银行", "", "", ""]
                if len(row) != 9:
                    icbc_ok = False
                    log(f"    列数错误: {old_name} 行 {row[0]} 有 {len(row)} 列 (期望 9)")
                    break
                # 跨行标识 = "1"
                if row[4] != "1":
                    icbc_ok = False
                    log(f"    跨行标识错误: {old_name} 行 {row[0]} = {row[4]} (期望 1)")
                # 行名 = "工商银行"
                if row[5] != "工商银行":
                    icbc_ok = False
                    log(f"    行名错误: {old_name} 行 {row[0]} = {row[5]} (期望 工商银行)")
                # 前4列非空
                if not row[1] or not row[2]:
                    icbc_ok = False
                    log(f"    账户/户名为空: {old_name} 行 {row[0]}")
        check(icbc_ok, "工商银行 3→9 列映射正确 (跨行标识=1, 行名=工商银行)")

        # 验证建设银行跨行标识留空
        log("  ── 建设银行跨行标识验证 ──")
        ccb_units = [r for r in renamed_list if r[2] == "建设银行"]
        ccb_ok = True
        for new_name, old_name, bank_name, unit_name in ccb_units:
            filepath = os.path.join(BANK_DIR, old_name)
            rows = _read_ccb_rows(filepath)
            for row in rows:
                if len(row) != 9:
                    ccb_ok = False
                    continue
                # 跨行标识留空
                if row[4] != "":
                    ccb_ok = False
                    log(f"    建设银行跨行标识不应为1: {old_name} 行 {row[0]} = '{row[4]}'")
                    break
                # 行名 = "建设银行"
                if row[5] != "建设银行":
                    ccb_ok = False
                    log(f"    建设银行行名错误: {old_name} 行 {row[0]} = {row[5]} (期望 建设银行)")
                    break
        check(ccb_ok, "建设银行跨行标识留空, 行名=建设银行")

        # 验证合并文件中的 3→9 映射实际生效（通过打开合并文件检查）
        log("  ── 合并文件中工行列映射验证 ──")
        icbc_in_merged = 0
        merge_mapping_ok = True
        for m in merged:
            # 找出包含工商银行且为多银行的合并文件
            if "工商银行" not in m:
                continue
            fp = os.path.join(tmpdir, m)
            try:
                wb = openpyxl.load_workbook(fp)
                ws = wb.active
                headers = [str(ws.cell(1, c).value).strip() for c in range(1, 10)]
                expected_headers = ["序号", "账户", "户名", "金额", "跨行标识", "行名", "联行行号", "摘要", "备注"]
                if headers != expected_headers:
                    merge_mapping_ok = False
                    log(f"    表头错误: {m} 实际={headers}")
                    continue

                # 检查每个工商银行行的跨行标识
                icbc_found = 0
                for r in range(2, ws.max_row + 1):
                    bank_name = str(ws.cell(r, 6).value or "").strip()
                    cross_flag = str(ws.cell(r, 5).value or "").strip()
                    if bank_name == "工商银行":
                        icbc_found += 1
                        if cross_flag != "1":
                            merge_mapping_ok = False
                            log(f"    合并文件中工行跨行标识错误: {m} 行 {r} = '{cross_flag}' (期望 1)")
                icbc_in_merged += icbc_found
            except Exception as e:
                merge_mapping_ok = False
                log(f"    打开合并文件失败: {m} → {e}")

        check(merge_mapping_ok, "合并文件中工行 3→9 列映射正确")
        check(icbc_in_merged > 0, f"合并文件中工行数据行数: {icbc_in_merged} > 0")

        return merged
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


# ──────────────────────────────────────────────
# 测试 3: 匹配测试
# ──────────────────────────────────────────────
def test_match(merged_list):
    log("\n" + "=" * 60)
    log("【测试 3】匹配测试")
    log("=" * 60)

    matched = match_payroll_files(merged_list, PAYROLL_DIR)

    check(len(matched) == 97, f"匹配结果数: {len(matched)} (期望 97)")

    # 验证匹配成功数
    matched_count = sum(1 for _, fp, _ in matched if fp is not None)
    check(matched_count == 97, f"匹配成功数: {matched_count}/97 (期望 97/97)")

    # 验证每个匹配都有有效文件路径
    all_valid = True
    for merged_name, fp, unit_name in matched:
        if fp is None or not os.path.isfile(fp):
            all_valid = False
            log(f"    无效文件路径: {unit_name} → {fp}")
    check(all_valid, "所有匹配的文件路径有效")

    # 验证 signed_ 优先
    log("  ── signed_ 优先验证 ──")
    signed_ok = True
    signed_count = 0
    for merged_name, fp, unit_name in matched:
        basename = os.path.basename(fp)
        if basename.startswith("signed_"):
            signed_count += 1
        # 检查是否有同时存在 signed_ 和 非 signed_ 版本时，优先选了 signed_
        # 在工资表目录中查找是否存在 signed_ 版本
        unit_base = unit_name  # 已从合并文件名提取
        has_signed = any(
            f.startswith("signed_" + unit_base) or
            (f.startswith("signed_") and unit_base in f)
            for f in os.listdir(PAYROLL_DIR)
        )
        if has_signed and not basename.startswith("signed_"):
            # 但需要确保确实没有 signed_ 匹配失败的情况
            # 先尝试一下是否应该是 signed_
            payroll_files = os.listdir(PAYROLL_DIR)
            signed_for_unit = [f for f in payroll_files
                               if f.startswith("signed_") and unit_base in f
                               and "汇总表" not in f and "验证" not in f]
            if signed_for_unit:
                signed_ok = False
                log(f"    应选 signed_ 但未选: {unit_name} → {basename}")
    check(signed_ok, "signed_ 版本优先选择")
    check(signed_count > 0, f"选择了 {signed_count} 个 signed_ 文件")

    # 验证汇总表被排除
    for _, fp, unit_name in matched:
        basename = os.path.basename(fp)
        if "汇总表" in basename or "验证" in basename:
            log(f"    错误: 匹配到汇总表/验证文件: {unit_name} → {basename}")
            check(False, "汇总表/验证文件被排除")
            break

    # 验证 unmatched 的细节
    unmatched = [(m, u) for m, f, u in matched if f is None]
    check(len(unmatched) == 0, f"无未匹配单位: {len(unmatched)} 个")

    return matched


# ──────────────────────────────────────────────
# 测试 4: 排序测试
# ──────────────────────────────────────────────
def test_sort(merged_list, matched_pairs):
    log("\n" + "=" * 60)
    log("【测试 4】排序测试")
    log("=" * 60)

    # 验证 merged_list 已排序
    is_sorted = all(merged_list[i] <= merged_list[i + 1] for i in range(len(merged_list) - 1))
    check(is_sorted, "merged_list 按文件名升序排列")

    # 验证 matched_pairs 顺序与 merged_list 一致
    order_ok = all(
        matched_pairs[i][0] == merged_list[i]
        for i in range(len(merged_list))
    )
    check(order_ok, "matched_pairs 顺序与 merged_list 一致")

    # 验证文件名排序符合 Windows 默认升序（逐字符比较）
    # Windows 排序规则: 数字<字母<中文, 逐字符比较
    sorted_manual = sorted(merged_list)
    order_match = merged_list == sorted_manual
    check(order_match, "排序结果与 Python 默认排序一致 (Windows 默认升序)")

    # 显示排序结果的前几个和后几个
    log(f"  首文件: {merged_list[0]}")
    log(f"  尾文件: {merged_list[-1]}")


# ──────────────────────────────────────────────
# 测试 5: 语法检查
# ──────────────────────────────────────────────
def test_syntax():
    log("\n" + "=" * 60)
    log("【测试 5】语法检查")
    log("=" * 60)

    import ast
    try:
        with open("/home/ubuntu/coding/seqprint/batchprint_gui.py", "r") as f:
            ast.parse(f.read())
        check(True, "batchprint_gui.py 语法检查通过")
    except SyntaxError as e:
        check(False, f"batchprint_gui.py 语法错误: {e}")


# ──────────────────────────────────────────────
# 主流程
# ──────────────────────────────────────────────
def main():
    log("=" * 60)
    log("batchprint_gui.py 集成测试")
    log(f"银行报盘: {BANK_DIR}")
    log(f"工资表:   {PAYROLL_DIR}")
    log("=" * 60)

    # 测试 5: 语法检查（最基础，先做）
    test_syntax()
    if HAS_FAILURE:
        log("\n⚠ 语法检查失败，终止后续测试")
        sys.exit(1)

    # 测试 1: 改名
    renamed = test_rename()
    if renamed is None:
        log("\n⚠ 改名测试异常终止")
        sys.exit(1)

    # 测试 2: 合并
    merged = test_merge(renamed)
    if merged is None:
        log("\n⚠ 合并测试异常终止")
        sys.exit(1)

    # 测试 3: 匹配
    matched = test_match(merged)

    # 测试 4: 排序
    test_sort(merged, matched)

    # ── 最终结果 ──
    log("\n" + "=" * 60)
    if HAS_FAILURE:
        log("❌ 存在测试失败！")
        sys.exit(1)
    else:
        log("✅ 全部测试通过！")
        log("=" * 60)


if __name__ == "__main__":
    main()
